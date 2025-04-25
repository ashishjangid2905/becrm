import os
import fiscalyear
from io import BytesIO
from pathlib import Path

# Import from django modules
from django.db.models import Q
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404

# Import from app model, utils permission etc
from .utils import (
    STATE_CHOICE,
    CATEGORY,
    REPORT_TYPE,
)
from invoice.models import proforma, BillerVariable, convertedPI, orderList

# Import from third party modules

from num2words import num2words

# reportLab Module
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    TableStyle,
    Image as Im,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

# pypdf for read and write
from pypdf import PdfReader, PdfWriter

# Openpyxl Module
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
)

from openpyxl.utils import get_column_letter

# funtion to use get roman no
def int_to_roman(num):
    val = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    syms = ["m", "cm", "d", "cd", "c", "xc", "l", "xl", "x", "ix", "v", "iv", "i"]

    roman_num = ""
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman_num += syms[i]
            num -= val[i]
        i += 1
    return roman_num

# function to format month value
def format_month(value):
    try:
        if value == None or value == "":
            return ""
        else:
            return datetime.strptime(value, "%Y-%m").strftime("%b'%y")
    except ValueError:
        return value

# Calculate total number of month
def total_month(from_month, to_month):
    try:
        from_m = datetime.strptime(from_month, "%Y-%m")
        to_m = datetime.strptime(to_month, "%Y-%m")
        return (to_m.year - from_m.year) * 12 + to_m.month - from_m.month + 1
    except (ValueError, TypeError):
        return 0


def multiply(value, arg):
    try:
        return value * arg
    except (ValueError, TypeError):
        return 0


# function to use categorise pi sales value by product
def sale_category(pi):
    online_sale = 0
    offline_sale = 0
    domestic_sale = 0
    unique_lumpsum_amt_online = set()  # To track unique lumpsum amounts
    unique_lumpsum_amt_offline = set()  # To track unique lumpsum amounts
    unique_lumpsum_amt_domestic = set()  # To track unique lumpsum amounts

    pi = proforma.objects.get(pk=pi)

    for order in pi.orderlist.all():
        if not order.is_lumpsum:
            if order.category == "online" and order.report_type == "online":
                online_sale += order.total_price
            elif order.category == "offline" and order.report_type != "domestic":
                offline_sale += order.total_price
            elif order.category == "offline" and order.report_type == "domestic":
                domestic_sale += order.total_price
        else:
            if order.category == "online" and order.report_type == "online":
                unique_lumpsum_amt_online.add(order.lumpsum_amt)
            elif order.category == "offline" and order.report_type != "domestic":
                unique_lumpsum_amt_offline.add(order.lumpsum_amt)
            elif order.category == "offline" and order.report_type == "domestic":
                unique_lumpsum_amt_domestic.add(order.lumpsum_amt)

    online_sale += sum(unique_lumpsum_amt_online)
    offline_sale += sum(unique_lumpsum_amt_offline)
    domestic_sale += sum(unique_lumpsum_amt_domestic)

    sale_category = {
        "online_sale": online_sale,
        "offline_sale": offline_sale,
        "domestic_sale": domestic_sale,
    }
    return sale_category

# calulate total pi value
def total_order_value(pi):

    pi = proforma.objects.get(pk=pi)

    total_sum = 0
    unique_lumpsum_amt = set()  # To track unique lumpsum amounts
    for order in pi.orderlist.all():
        if order.is_lumpsum and order.lumpsum_amt:
            unique_lumpsum_amt.add(order.lumpsum_amt)
        elif not order.is_lumpsum:
            total_sum += order.total_price

    total_sum += sum(unique_lumpsum_amt)
    return total_sum

# calculate total pi value including tax as tax criterias
def total_pi_value_inc_tax(pi):

    pi_instance = proforma.objects.get(pk=pi)
    total_value = total_order_value(pi)

    if not pi_instance.bank.biller_id.biller_gstin or pi_instance.is_sez:
        cgst = 0
        sgst = 0
        igst = 0
        total_inc_tax = total_value
    else:
        if str(pi_instance.state) == pi_instance.bank.biller_id.biller_gstin[0:2]:
            cgst = total_value * 0.09
            sgst = total_value * 0.09
            igst = 0
            total_inc_tax = total_value * 1.18
        elif str(pi_instance.state) == "500":
            cgst = 0
            sgst = 0
            igst = 0
            total_inc_tax = total_value
        else:
            cgst = 0
            sgst = 0
            igst = total_value * 0.18
            total_inc_tax = total_value * 1.18

    return f"{round(total_inc_tax,0):.0f}"

# filter pi orders according to Is_lumpsum
def filter_by_lumpsum(queryset, is_lumpsum):
    return queryset.filter(is_lumpsum=is_lumpsum)

# calulate total lumpsum amount in PI if available
def total_lumpsums(pi):
    pi_instance = proforma.objects.get(pk=pi)
    unique_lumpsum_amt = set()  # To track unique lumpsum amounts
    total_lumpsum_amt = 0
    for order in pi_instance.orderlist.all():
        if order.is_lumpsum and order.lumpsum_amt:
            unique_lumpsum_amt.add(order.lumpsum_amt)
    total_lumpsum_amt += sum(unique_lumpsum_amt)
    return total_lumpsum_amt


def split(value, delimiter):
    """Splits the string by the given delimiter."""
    return value.split(delimiter)

# calculate total due amount as per pi total value and reeived value
def total_dues(pi):
    pi = proforma.objects.get(pk=pi)
    total_amt = int(total_pi_value_inc_tax(pi))
    convertedpi = getattr(proforma, "convertedpi", None)
    if not convertedpi:
        return total_amt
    payment1_amt = int(getattr(convertedpi, "payment1_amt", 0) or 0)
    payment2_amt = int(getattr(convertedpi, "payment2_amt", 0) or 0)
    payment3_amt = int(getattr(convertedpi, "payment3_amt", 0) or 0)
    total_receive = payment1_amt + payment2_amt + payment3_amt
    total_due = total_amt - total_receive
    return int(total_due)

# get billers additional active variables
def get_biller_variable(biller_dtl, variable_name):
    today = timezone.now().date()

    filters = {
        "from_date__lte": today,
        "biller_id": biller_dtl,
        "variable_name": variable_name,
    }

    variable = BillerVariable.objects.filter(
        Q(to_date__gte=today) | Q(to_date__isnull=True), **filters
    ).last()

    return variable.variable_value if variable else None

# get new pi number
def pi_number(biller_id):
    fiscalyear.START_MONTH = 4
    fy = str(fiscalyear.FiscalYear.current())[-2:]
    py = str(int(fy) - 1)

    pi_tag = get_biller_variable(biller_id, "pi_tag")
    pi_format = get_biller_variable(biller_id, "pi_format")

    if not isinstance(pi_format, str):
        raise ValueError(f"Expected pi_format to be a string, got {type(pi_format)}")

    search_prefix = pi_format.format(py=py, fy=fy, tag=pi_tag, num=0).rstrip("0")

    last_pi = (
        proforma.objects.filter(pi_no__startswith=search_prefix)
        .order_by("pi_no")
        .last()
    )

    if last_pi:
        try:
            last_no = int(last_pi.pi_no.split(search_prefix)[-1])
        except:
            last_no = 0
        new_no = last_no + 1
    else:
        new_no = 1

    pi_no = pi_format.format(py=py, fy=fy, tag=pi_tag, num=new_no)

    return pi_no

# get current fisal year 
def current_fy(date=None):
    today = timezone.now().date()
    if date:
        today = datetime.strptime(date, "%Y-%m-%d").date()
    year = today.year

    if today.month >= 4:
        start_year = year
        end_year = year + 1
    else:
        start_year = year - 1
        end_year = year

    return f"{start_year}-{end_year}"

# get formatted invoice no as per current fiscal year
def get_invoice_no(biller_id, invoice_tag, invoice_format, new_no=None):
    fiscalyear.START_MONTH = 4
    fy = str(fiscalyear.FiscalYear.current())[-2:]
    py = str(int(fy) - 1)

    if not isinstance(invoice_format, str):
        raise ValueError(
            f"Expected pi_format to be a string, got {type(invoice_format)}"
        )

    search_prefix = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=0).rstrip(
        "0"
    )

    if not new_no:
        last_invoice = (
            convertedPI.objects.filter(
                invoice_no__startswith=search_prefix, pi_id__bank__biller_id=biller_id
            )
            .order_by("pi_no")
            .last()
        )

        if last_invoice:
            try:
                last_no = int(last_invoice.pi_no.split(search_prefix)[-1])
            except:
                last_no = 0
            new_no = last_no + 1
        else:
            new_no = 1

    invoice_no = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=new_no)

    return invoice_no

# get formatted invoice no as per invoice date fiscal year
def get_invoice_no_from_date(
    biller_id, invoice_tag, invoice_format, invoice_date, new_no=None
):
    fiscalYear = current_fy(invoice_date)
    fy = fiscalYear.split("-")[1][-2:]
    py = fiscalYear.split("-")[0][-2:]

    if not isinstance(invoice_format, str):
        raise ValueError(
            f"Expected pi_format to be a string, got {type(invoice_format)}"
        )

    search_prefix = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=0).rstrip(
        "0"
    )

    if not new_no:
        last_invoice = (
            convertedPI.objects.filter(
                invoice_no__startswith=search_prefix, pi_id__bank__biller_id=biller_id
            )
            .order_by("pi_no")
            .last()
        )

        if last_invoice:
            try:
                last_no = int(last_invoice.pi_no.split(search_prefix)[-1])
            except:
                last_no = 0
            new_no = last_no + 1
        else:
            new_no = 1

    invoice_no = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=new_no)

    return invoice_no

# export xls file
def exportInvoicelist(invoices):

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "S.N",
        "Team Member",
        "Company Name",
        "GSTIN",
        "Address",
        "State",
        "Country",
        "PI No",
        "PI Date",
        "Contact Person Name",
        "Email",
        "Contact Number",
        "Bank",
        "A/C No",
        "Beneficry Name",
        "Payment Status",
        "1st Payment",
        "1st Payment Date",
        "2nd Payment",
        "2nd Payment Date",
        "3rd Payment",
        "3rd Payment Date",
        "Amount",
        "Amount (inc. tax)",
        "Total Received",
        "Is Generated",
        "Invoice No",
        "Invoice Date",
    ]

    ws.append(headers)

    header_fill = PatternFill(
        fill_type="solid", start_color="0099CCFF", end_color="0099CCFF"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="00000000")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    s = 1

    for invoice in invoices:
        total_received = (
            int(
                invoice.convertedpi.payment1_amt
                if invoice.convertedpi.payment1_amt != None
                else 0
            )
            + int(
                invoice.convertedpi.payment2_amt
                if invoice.convertedpi.payment2_amt != None
                else 0
            )
            + int(
                invoice.convertedpi.payment3_amt
                if invoice.convertedpi.payment3_amt != None
                else 0
            )
        )

        row = [
            s,
            invoice.user_name,
            invoice.company_name,
            invoice.gstin,
            invoice.address,
            dict(STATE_CHOICE).get(int(invoice.state)),
            invoice.country,
            invoice.pi_no,
            invoice.pi_date,
            invoice.requistioner,
            invoice.email_id,
            invoice.contact,
            invoice.bank.bank_name,
            invoice.bank.ac_no,
            invoice.bank.bnf_name,
            invoice.convertedpi.payment_status,
            invoice.convertedpi.payment1_amt,
            invoice.convertedpi.payment1_date,
            invoice.convertedpi.payment2_amt,
            invoice.convertedpi.payment2_date,
            invoice.convertedpi.payment3_amt,
            invoice.convertedpi.payment3_date,
            total_order_value(invoice.id),
            total_pi_value_inc_tax(invoice.id),
            total_received,
            invoice.convertedpi.is_taxInvoice,
            invoice.convertedpi.invoice_no,
            invoice.convertedpi.invoice_date,
        ]

        ws.append(row)
        s += 1

    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    today = timezone.now().date()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="invoice-list-{today}.xlsx"'
    )
    wb.save(response)
    return response

# download pdf file of invoice
def pdf_PI(pi_id):
    pi = get_object_or_404(proforma, pk=pi_id)
    orders = orderList.objects.filter(proforma_id=pi)

    reg_address = pi.bank.biller_id.get_reg_full_address()
    corp_address = pi.bank.biller_id.get_corp_full_address()

    net_total = total_order_value(pi.id)
    lumpsumAmt = total_lumpsums(pi.id)

    if pi.currency == "inr":
        curr = "₹"
    else:
        curr = "$"

    if pi.is_sez:
        cgst = 0
        sgst = 0
        igst = 0
        total_inc_tax = net_total
    else:
        if pi.bank.biller_id.biller_gstin:
            if str(pi.state) == pi.bank.biller_id.biller_gstin[0:2]:
                cgst = net_total * 0.09
                sgst = net_total * 0.09
                igst = 0
                total_inc_tax = net_total * 1.18
            elif pi.state == "500":  # for Foreign Clients
                cgst = 0
                sgst = 0
                igst = 0
                total_inc_tax = net_total
            else:
                cgst = 0
                sgst = 0
                igst = net_total * 0.18
                total_inc_tax = net_total * 1.18
        else:
            cgst = 0
            sgst = 0
            igst = 0
            total_inc_tax = net_total

    roundOff = total_inc_tax - round(total_inc_tax, 0)

    if pi.currency == "inr":
        total_val_words = f"Rs. {num2words(round(total_inc_tax,0), lang='en-IN').replace(',', '').title()} Only"
    else:
        total_val_words = f"Usd {num2words(round(total_inc_tax,0), lang='en-IN').replace(',', '').title()} Only"

    basedir = Path(__file__).resolve().parent.parent

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=0 * inch,
        rightMargin=0 * inch,
        topMargin=0.2 * inch,
        bottomMargin=3 * inch,
    )

    pdfmetrics.registerFont(
        TTFont(
            "Montserrat-Light",
            os.path.join(basedir, "static/becrm/fonts/Montserrat-Light.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Montserrat-Regular",
            os.path.join(basedir, "static/becrm/fonts/Montserrat-Regular.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Montserrat-Medium",
            os.path.join(basedir, "static/becrm/fonts/Montserrat-Medium.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Montserrat-Bold",
            os.path.join(basedir, "static/becrm/fonts/Montserrat-Bold.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Quicksand-Bold",
            os.path.join(basedir, "static/becrm/fonts/Quicksand-Bold.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Quicksand-Light",
            os.path.join(basedir, "static/becrm/fonts/Quicksand-Light.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Quicksand-Medium",
            os.path.join(basedir, "static/becrm/fonts/Quicksand-Medium.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Quicksand-Regular",
            os.path.join(basedir, "static/becrm/fonts/Quicksand-VariableFont_wght.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Quicksand-SemiBold",
            os.path.join(basedir, "static/becrm/fonts/Quicksand-SemiBold.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Roboto-Medium",
            os.path.join(basedir, "static/becrm/fonts/Roboto-Medium.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Roboto-Regular",
            os.path.join(basedir, "static/becrm/fonts/Roboto-Regular.ttf"),
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Roboto-Bold", os.path.join(basedir, "static/becrm/fonts/Roboto-Bold.ttf")
        )
    )
    pdfmetrics.registerFont(
        TTFont(
            "Roboto-Light", os.path.join(basedir, "static/becrm/fonts/Roboto-Light.ttf")
        )
    )

    blue_font = ParagraphStyle(
        name="BlueStyle",
        fontSize=24,
        alignment=0,
        fontName="Helvetica-Bold",
        leading=20,
        underlineWidth=2,
    )
    black_font = ParagraphStyle(
        name="BlueStyle", fontSize=12, alignment=0, fontName="Montserrat-Light"
    )

    font_xxs = ParagraphStyle(
        name="font_xxs",
        fontSize=8,
        fontName="Quicksand-bold",
        alignment=0,
        leading=8,
        textColor=colors.HexColor("#ffffff"),
    )
    font_xs = ParagraphStyle(
        name="font_xs",
        fontSize=7,
        fontName="Quicksand-Bold",
        alignment=0,
        leading=7,
        textColor=colors.HexColor("#ffffff"),
    )
    font_s = ParagraphStyle(
        name="font_s", fontSize=10, fontName="Quicksand-light", alignment=0
    )

    brand_text = f"{pi.bank.biller_id.brand_name}"
    brand_name = Paragraph(f'<font color="#3182d9">{brand_text}</font>', blue_font)

    if pi.bank.biller_id.biller_name != pi.bank.biller_id.brand_name:
        founder_text = f"{pi.bank.biller_id.biller_name}"
        founder_name = Paragraph(
            f'<font color="#000000" >founded by {founder_text}</font>', black_font
        )
    else:
        founder_name = ""

    imagepath = os.path.join(basedir, "static/becrm/image/pi_back.png")
    logo = Im(imagepath, 8.25 * inch, 3 * inch)

    biller_font = ParagraphStyle(
        name="BlueStyle",
        fontSize=11,
        alignment=0,
        textColor="white",
        fontName="Roboto-Bold",
    )

    if corp_address:
        corpAddress = Paragraph(
            f"<strong>Corporate Office: </strong><font>{corp_address}</font>", font_xs
        )
    else:
        corpAddress = ""
    regAddress = Paragraph(f"<b>Reg. Office: </b><font>{reg_address}</font>", font_xs)
    gstin_text = (
        f"<b>GSTIN: </b><font>{pi.bank.biller_id.biller_gstin}</font>"
        if corp_address
        else ""
    )

    biller_Name = Paragraph(
        f"<b>{pi.bank.biller_id.biller_name} | {gstin_text}</b>", biller_font
    )

    address = Paragraph(f"<b>Address: </b><font>{pi.address}</font>", font_xxs)

    email = Paragraph(f"<b>E: </b><font>{pi.email_id}</font>", font_xxs)
    contact = Paragraph(f"<b>M: </b><font>{pi.contact}</font>", font_xxs)

    piDate = pi.pi_date.strftime("%d-%b-%y")
    poDate = pi.po_date.strftime("%d-%b-%y") if pi.po_date else ""

    if hasattr(pi, "convertedpi"):
        invoice_type = (
            "TAX INVOICE" if pi.convertedpi.is_taxInvoice else "PRO FORMA INVOICE"
        )
    else:
        invoice_type = "PRO FORMA INVOICE"

    invoice_date = (
        pi.convertedpi.invoice_date.strftime("%d-%b-%y")
        if invoice_type == "TAX INVOICE"
        else ""
    )

    data = [
        [brand_name, "", "", "", "", "", "", "", "", "", ""],
        [founder_name, "", "", "", "", "", "", "", "", "", invoice_type],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", logo, "", "", "", "", ""],
        [
            "Issued by:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f"{curr} {round(total_inc_tax,0):.2f}",
            "",
            "",
        ],
        [
            biller_Name,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            regAddress,
            "",
            "",
            corpAddress,
            "",
            "",
            "",
            "",
            "Total Payable Amount",
            "",
            "",
        ],
        ["", "", "", "", "", "", "", "", "", "", ""],
        [
            "Customer Details:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "VENDOR CODE:" if pi.vendor_code else "",
            pi.vendor_code,
            "",
        ],
        [
            pi.company_name.upper(),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "LUT NO:" if pi.lut_no else "",
            pi.lut_no,
            "",
        ],
        [
            f"GSTIN: {pi.gstin}",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "PO DATE:" if pi.po_date else "",
            poDate if pi.po_date else "",
            "",
        ],
        [
            address,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "PO NUMBER:" if pi.po_no else "",
            pi.po_no if pi.po_no else "",
            "",
        ],
        ["", "", "", "", "", "", "", "", "PI DATE:", piDate, ""],
        [email, "", "", "", "", "", "", "", "PI NUMBER:", pi.pi_no, ""],
        [
            contact,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Invoice Date:" if invoice_type == "TAX INVOICE" else "",
            invoice_date if invoice_type == "TAX INVOICE" else "",
            "",
        ],
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Invoice No.:" if invoice_type == "TAX INVOICE" else "",
            pi.convertedpi.invoice_no if invoice_type == "TAX INVOICE" else "",
            "",
        ],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        [
            "",
            "S.N",
            "ITEM DESCRIPTION",
            "",
            "",
            "",
            "",
            "RATE",
            "",
            f"TOTAL ({curr})",
            "",
        ],
        ["", "", "", "", "", "", "", "", "", "", ""],
    ]

    row_heights = [0.4 * inch, 0.3 * inch] + 3 * [0.1 * inch] + 15 * [0.2 * inch]
    col_widths = 11 * [0.7 * inch]

    table = Table(data, col_widths, row_heights)

    style = TableStyle(
        [
            ("SPAN", (0, 0), (5, 0)),
            ("FONT", (0, 1), (6, 1), "Montserrat-Regular", 13),
            ("SPAN", (0, 1), (5, 1)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("VALIGN", (7, 1), (10, 1), "BOTTOM"),
            ("TEXTCOLOR", (7, 1), (10, 1), colors.HexColor("#545454")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (5, 3), (5, 3), "CENTER"),
            ("ALIGN", (7, 1), (10, 4), "RIGHT"),
            ("FONT", (7, 1), (10, 3), "Montserrat-Regular", 22),
            ("FONT", (7, 3), (10, 4), "Roboto-Medium", 16),
            ("TEXTCOLOR", (7, 3), (10, 4), colors.HexColor("#545454")),
            (
                "VALIGN",
                (0, 4),
                (6, 4),
                "MIDDLE",
            ),
            ("FONT", (0, 4), (6, 4), "Montserrat-Medium", 8),
            ("TEXTCOLOR", (0, 3), (6, 4), colors.HexColor("#ffffff")),
            ("SPAN", (8, 4), (10, 4)),
            ("SPAN", (8, 6), (10, 6)),
            ("ALIGN", (7, 4), (10, 6), "CENTER"),
            ("VALIGN", (7, 6), (10, 6), "TOP"),
            ("FONT", (7, 6), (10, 6), "Montserrat-Light", 9),
            ("ALIGN", (7, 3), (9, 3), "CENTER"),
            ("SPAN", (0, 5), (5, 5)),
            ("SPAN", (0, 6), (2, 6)),
            ("SPAN", (3, 6), (5, 6)),
            ("FONT", (0, 8), (10, 8), "Montserrat-Regular", 9),
            ("FONT", (0, 9), (4, 9), "Roboto-Bold", 13),
            ("FONT", (0, 10), (4, 10), "Roboto-Regular", 10),
            ("SPAN", (0, 11), (5, 11)),
            ("SPAN", (0, 13), (6, 13)),
            ("SPAN", (0, 14), (5, 14)),
            ("ALIGN", (8, 8), (8, 15), "RIGHT"),
            ("FONT", (8, 8), (8, 15), "Quicksand-Bold", 9),
            ("FONT", (9, 8), (9, 15), "Quicksand-Medium", 9),
            ("LINEAFTER", (6, 9), (6, 13), 2, colors.white),
            ("TEXTCOLOR", (0, 7), (10, 15), colors.HexColor("#ffffff")),
            ("FONT", (0, 18), (10, 18), "Quicksand-Bold", 11),
            ("ALIGN", (7, 18), (9, 19), "RIGHT"),
        ]
    )

    table.setStyle(style)

    nr = 0
    s = 1
    order_list = []
    lumpsumRow = 0

    for order in filter_by_lumpsum(orders, True):
        if order.is_lumpsum:
            cat = dict(CATEGORY).get(str(order.category))
            report = dict(REPORT_TYPE).get(str(order.report_type))
            product = order.product
            from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
            to_month = datetime.strptime(order.to_month, "%Y-%m").strftime("%b'%y")
            period = (
                f"{from_month} - {to_month}"
                if from_month != to_month
                else f"{from_month}"
            )

            if pi.currency == "inr":
                unitPrice = f"Lumpsum"
                totalPrice = f"{lumpsumAmt:.2f}"
            else:
                unitPrice = f"Lumpsum"
                totalPrice = f"{lumpsumAmt:.2f}"

            sac = f"SAC: 998399"

        from_m = datetime.strptime(order.from_month, "%Y-%m")
        to_m = datetime.strptime(order.to_month, "%Y-%m")

        no_months = (to_m.year - from_m.year) * 12 + to_m.month - from_m.month + 1

        total_m = f"{no_months} months"

        orderDes = Paragraph(
            f"<font>{report} | {product} | {"Period" if order.category =="offline" else "Validity"}: {period if order.category =="offline" else total_m}</font>",
            font_s,
        )

        order_list.extend(
            [
                [
                    f"{int_to_roman(s)})",
                    f"{cat} | {sac}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ],
                ["", orderDes, "", "", "", "", "", "", ""],
            ]
        )
        nr += 2
        s += 1
        lumpsumRow += 2

    for order in filter_by_lumpsum(orders, False):
        cat = dict(CATEGORY).get(str(order.category))
        report = dict(REPORT_TYPE).get(str(order.report_type))
        product = order.product
        from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
        to_month = datetime.strptime(order.to_month, "%Y-%m").strftime("%b'%y")

        period = f"{from_month} - {to_month}"

        unitPrice = f"{curr} {order.unit_price}"
        totalPrice = f"{order.total_price:.2f}"

        sac = f"SAC: 998399"

        from_m = datetime.strptime(order.from_month, "%Y-%m")
        to_m = datetime.strptime(order.to_month, "%Y-%m")

        no_months = (to_m.year - from_m.year) * 12 + to_m.month - from_m.month + 1

        total_m = f"Total: {no_months}"

        orderDes = Paragraph(
            f"<font>{report} | {product} | {"Period" if order.category =="offline" else "Validity"}: {period if order.category =="offline" else total_m}</font>",
            font_s,
        )

        order_list.extend(
            [
                [
                    f"{int_to_roman(s)})",
                    f"{cat} | {sac}",
                    "",
                    "",
                    "",
                    "",
                    unitPrice,
                    "",
                    totalPrice,
                ],
                ["", orderDes, "", "", "", "", "", "", ""],
            ]
        )
        nr += 2
        s += 1

    orderStyle = []

    if lumpsumRow > 0:
        for r in range(0, lumpsumRow, 2):
            end_row = min(lumpsumRow - 1, r + 1)
            order_list[0][6] = "Lumpsum"
            order_list[0][8] = f"{lumpsumAmt:.2f}"
            for i in range(r + 1, end_row + 1):
                order_list[i][6] = ""
                order_list[i][8] = ""
            orderStyle.extend(
                [
                    ("SPAN", (6, r), (6, lumpsumRow - 1)),
                    ("SPAN", (8, r), (8, lumpsumRow - 1)),
                ]
            )

    if nr > lumpsumRow:
        for r in range(lumpsumRow, nr, 2):
            orderStyle.extend(
                [
                    ("SPAN", (6, r), (6, r + 1)),
                    ("SPAN", (8, r), (8, r + 1)),
                ]
            )

    for r in range(0, nr, 2):
        orderStyle.extend(
            [
                ("LINEABOVE", (0, 0), (-1, 0), 2, colors.HexColor("#c63545")),
                ("SPAN", (1, r), (5, r)),
                ("SPAN", (1, r + 1), (5, r + 1)),
                ("FONT", (1, r), (5, r), "Quicksand-Bold", 10),
                ("FONT", (1, r + 1), (5, r + 1), "Quicksand-Medium", 10),
                ("FONT", (6, r), (8, -1), "Quicksand-Medium", 10),
                ("ALIGN", (6, r), (8, -1), "RIGHT"),
                ("RIGHTPADDING", (6, r), (6, -1), -10),
            ]
        )

    orderStyle.append(
        (
            "ROWBACKGROUNDS",
            (0, 0),
            (5, -1),
            [
                colors.white,
                colors.white,
                colors.HexColor("#f1f1f1"),
                colors.HexColor("#f1f1f1"),
            ],
        )
    )

    order_table = Table(order_list, col_widths)

    order_table.setStyle(orderStyle)

    pdf_canvas = canvas.Canvas(buffer, pagesize=A4)

    # Draw something on the canvas (example)
    def add_canvas(canvas, doc):

        canvas.setLineWidth(1)  # Set the border width

        # You can also reset the font color if needed for further text
        canvas.setFillColor(
            colors.HexColor("#004aad")
        )  # Reset to black or any other color after drawing
        canvas.setStrokeColor(colors.white)  # Set stroke color for the border
        canvas.setLineWidth(1)
        canvas.roundRect(360, 165, 200, 95, 10, fill=1)

        canvas.setFillColor(
            colors.HexColor("#3182d9")
        )  # Reset to black or any other color after drawing
        canvas.rect(30, 101, 150, 30, fill=1)
        canvas.rect(180, 101, 160, 30, fill=1)
        canvas.rect(340, 101, 120, 30, fill=1)
        canvas.rect(460, 101, 100, 30, fill=1)

        canvas.setFillColor(
            colors.HexColor("#38b6ff")
        )  # Reset to black or any other color after drawing
        canvas.rect(30, 70, 150, 30, fill=1)
        canvas.rect(180, 70, 160, 30, fill=1)
        canvas.rect(340, 70, 120, 30, fill=1)
        canvas.rect(460, 70, 100, 30, fill=1)

        canvas.rect(30, 39, 530, 30, fill=1)

    def add_text(canvas, doc):
        canvas.setFont("Quicksand-Medium", 10)
        canvas.setFillColor(
            colors.HexColor("#ffffff")
        )  # Set the font color to a shade of blue

        # Set font color for the tax strings
        canvas.drawRightString(440, 245, f"AMOUNT:")
        canvas.drawRightString(530, 245, f"{curr} {net_total:.2f}")
        canvas.drawRightString(440, 230, "IGST (18%):")
        canvas.drawRightString(530, 230, f"{curr} {igst:.2f}")
        canvas.drawRightString(440, 215, "CGST (9%):")  # Use different y-coordinates
        canvas.drawRightString(
            530, 215, f"{curr} {cgst:.2f}"
        )  # Use different y-coordinates
        canvas.drawRightString(440, 200, "SGST (9%):")  # Use different y-coordinates
        canvas.drawRightString(
            530, 200, f"{curr} {sgst:.2f}"
        )  # Use different y-coordinates

        canvas.setFont("Quicksand-Medium", 8)
        canvas.drawRightString(
            440, 188, "rounded off (+/-):"
        )  # Use different y-coordinates
        canvas.drawRightString(
            530, 188, f"{curr} {roundOff:.2f}"
        )  # Use different y-coordinates
        canvas.setFont("Quicksand-Bold", 12)
        canvas.drawRightString(440, 173, f"TOTAL:")  # Use different y-coordinates
        canvas.drawRightString(
            530, 173, f"{curr} {round(total_inc_tax,0):.2f}"
        )  # Use different y-coordinates

        canvas.setFont("Quicksand-Regular", 11)
        canvas.setFillColor(
            colors.HexColor("#004aad")
        )  # Set the font color to a shade of blue
        canvas.drawRightString(
            350, 230, f"AMOUNT PAYABLE IN WORDS:"
        )  # Use different y-coordinates

        words = total_val_words.split(" ")
        line = []
        current_line = ""

        for word in words:
            test_line = current_line + word + " "

            if canvas.stringWidth(test_line, "Quicksand-Bold", 12) <= 250:
                current_line = test_line
            else:
                line.append(current_line.strip())
                current_line = word + " "
        if current_line:
            line.append(current_line.strip())

        line_height = 13

        canvas.setFont("Quicksand-Bold", 12)
        for i, line in enumerate(line):

            canvas.drawRightString(
                350, 215 - (i * line_height), f"{line}"
            )  # Use different y-coordinates

        canvas.setFont("Montserrat-Medium", 10)
        canvas.setFillColor(
            colors.HexColor("#3182d9")
        )  # Set the font color to a shade of blue
        canvas.drawRightString(
            540, 145, "Whether Tax is payable under REVERSE CHARGE: "
        )  # Use different y-coordinates
        canvas.setFont("Montserrat-Bold", 10)
        canvas.drawRightString(560, 145, "NO")  # Use different y-coordinates

        canvas.setFillColor(
            colors.HexColor("#ffffff")
        )  # Set the font color to a shade of blue
        canvas.setFont("Montserrat-Bold", 10)
        canvas.drawCentredString(105, 112, "Team Member")
        canvas.drawCentredString(260, 112, "Requisitioner")
        canvas.drawCentredString(400, 112, "Subscription Mode")
        canvas.drawCentredString(510, 112, "Payment Term")

        canvas.setFont("Montserrat-Medium", 9)
        canvas.drawCentredString(105, 80, f"{pi.user_name}")
        canvas.drawCentredString(260, 80, f"{pi.requistioner}")
        canvas.drawCentredString(400, 80, f"{pi.subscription.capitalize()}")
        canvas.drawCentredString(510, 80, f"{pi.payment_term.capitalize()}")

        email_H_width = canvas.stringWidth("Email: ", "Montserrat-Medium", 9)
        email_width = canvas.stringWidth(f"{pi.user_email}", "Montserrat-Bold", 9)
        contact_H_width = canvas.stringWidth(" | Contact: ", "Montserrat-Medium", 9)
        contact_width = canvas.stringWidth(f"{pi.user_contact}", "Montserrat-Bold", 9)
        web_H_width = canvas.stringWidth(" | Website: ", "Montserrat-Medium", 9)
        web_width = canvas.stringWidth("www.besmartexim.com", "Montserrat-Bold", 9)

        start_point = (
            295
            - (
                email_H_width
                + email_width
                + contact_H_width
                + contact_width
                + web_H_width
                + web_width
            )
            / 2
        )

        canvas.setFont("Montserrat-Medium", 9)
        canvas.drawString(start_point, 50, "Email: ")
        canvas.drawString(start_point + email_H_width + email_width, 50, " | Contact: ")
        canvas.drawString(
            start_point + email_H_width + email_width + contact_H_width + contact_width,
            50,
            " | Website: ",
        )
        canvas.setFont("Montserrat-Bold", 9)
        canvas.drawString(start_point + email_H_width, 50, f"{pi.user_email}")
        canvas.drawString(
            start_point + email_H_width + email_width + contact_H_width,
            50,
            f"{pi.user_contact}",
        )
        canvas.drawString(
            start_point
            + email_H_width
            + email_width
            + contact_H_width
            + contact_width
            + web_H_width,
            50,
            "www.besmartexim.com",
        )

        canvas.setFont("Montserrat-Regular", 7)
        canvas.setFillColor(colors.HexColor("#545454"))
        canvas.drawCentredString(300, 10, f"This is Computer generated {invoice_type}")

    def add_last_page_text(canvas):

        canvas.setFillColor(
            colors.HexColor("#004aad")
        )  # Reset to black or any other color after drawing
        canvas.setStrokeColor(colors.white)  # Set stroke color for the border
        canvas.setLineWidth(1)
        canvas.roundRect(30, 640, 290, 170, 10, fill=1)

        canvas.setFont("Montserrat-Medium", 11)
        canvas.setFillColor(colors.HexColor("#ffffff"))
        canvas.drawString(50, 780, "Kindly pay in favor of")
        if pi.bank.is_upi:
            canvas.drawRightString(145, 740, "Holder Name: ")
            canvas.drawRightString(145, 725, "UPI ID.: ")
            canvas.drawRightString(145, 710, "UPI No.: ")
        else:
            canvas.drawRightString(145, 740, "Bank Name: ")
            canvas.drawRightString(145, 725, "A/C No.: ")
            canvas.drawRightString(145, 710, "Branch Address: ")
            canvas.drawRightString(145, 660, "IFSC: ")
            if pi.country != "IN":
                canvas.drawRightString(145, 645, "SWIFT: ")
            canvas.setFont("Montserrat-Bold", 11)

        if pi.bank.is_upi:
            canvas.drawString(150, 740, f"{pi.bank.bnf_name}")
            canvas.drawString(150, 725, f"{pi.bank.upi_id}")
            canvas.drawString(150, 710, f"{pi.bank.upi_no}")
        else:
            canvas.drawString(50, 760, f"{pi.bank.bnf_name}")
            canvas.drawString(150, 740, f"{pi.bank.bank_name}")
            canvas.drawString(150, 725, f"{pi.bank.ac_no}")

            canvas.drawString(150, 660, f"{pi.bank.ifsc}")
            if pi.country != "IN":
                canvas.drawString(150, 645, f"{pi.bank.swift_code}")

        def wrap_string(string, font_size, wrap_length):

            # Wraping Branch Address
            string = string.split(" ")
            line = []
            current_line = ""

            for word in string:
                test_line = current_line + word + " "
                if (
                    canvas.stringWidth(test_line, "Montserrat-Bold", font_size)
                    <= wrap_length
                ):
                    current_line = test_line
                else:
                    line.append(current_line.strip())
                    current_line = word + " "
            if current_line:
                line.append(current_line.strip())

            return line

        if not pi.bank.is_upi:
            line = wrap_string(pi.bank.branch_address, 11, 150)
            line_height = 13
            for i, line in enumerate(line):
                canvas.drawString(150, 710 - (i * line_height), f"{line}")

        canvas.setFont("Montserrat-Bold", 11)
        canvas.setFillColor(colors.HexColor("#545454"))
        canvas.drawString(
            30,
            580,
            f"Other {pi.bank.biller_id.biller_name} Details for future reference",
        )
        canvas.setFillColor(colors.HexColor("#3182d9"))
        canvas.drawString(30, 500, f"Terms & Conditions")

        canvas.setFont("Montserrat-Regular", 11)
        canvas.setFillColor(colors.HexColor("#545454"))
        canvas.circle(50, 553, 2, stroke=1, fill=1)
        canvas.drawString(55, 550, f"PAN: {pi.bank.biller_id.biller_pan}")
        canvas.circle(50, 533, 2, stroke=1, fill=1)
        canvas.drawString(
            55,
            530,
            f"Udyam Registration Number: {pi.bank.biller_id.biller_msme if pi.bank.biller_id.biller_msme else 'N/A' }",
        )

        canvas.setFont("Montserrat-Regular", 9)
        canvas.setFillColor(colors.HexColor("#3182d9"))
        canvas.circle(50, 483, 2, stroke=1, fill=1)
        canvas.drawString(55, 480, "Payment has to be made 100% in advance")
        canvas.circle(50, 463, 2, stroke=1, fill=1)
        canvas.drawString(
            55, 460, f"Company Legal Name is {pi.bank.biller_id.biller_name}"
        )
        canvas.circle(50, 443, 2, stroke=1, fill=1)
        canvas.drawString(
            55,
            440,
            "Any changes in the order will attract cost once order is sent based on PI approval by the client",
        )
        canvas.circle(50, 423, 2, stroke=1, fill=1)
        canvas.drawString(55, 420, f"18% GST will be applicable for all transactions")
        canvas.circle(50, 403, 2, stroke=1, fill=1)
        canvas.drawString(
            55,
            400,
            f"E-Invoice is applicable w.e.f. 1st August 2023 and hence once filed, no changes cannot be made after 24 hrs.",
        )

    elements = []
    elements.append(table)
    elements.append(order_table)

    buffer.seek(0)

    doc.build(elements)

    original_pdf = PdfReader(buffer)
    total_pages = len(original_pdf.pages)

    buffer_2 = BytesIO()

    pdf_writer = PdfWriter()

    for page_num in range(total_pages):
        pdf_page = original_pdf.pages[page_num]
        packet1 = BytesIO()
        page_canvas = canvas.Canvas(packet1, pagesize=A4)
        page_number_text = f"Page {page_num + 1} of {total_pages+1}"
        page_canvas.setFont("Montserrat-Regular", 8)
        page_canvas.setFillColor(colors.HexColor("#545454"))
        page_canvas.drawString(510, 10, page_number_text)
        page_canvas.save()
        packet1.seek(0)
        page_with_number = PdfReader(packet1)
        pdf_page.merge_page(page_with_number.pages[0])

        if page_num == total_pages - 1:
            packet = BytesIO()
            pdf_canvas1 = canvas.Canvas(packet, pagesize=A4)
            add_canvas(pdf_canvas1, pdf_page)
            add_text(pdf_canvas1, pdf_page)

            pdf_canvas1.save()

            packet.seek(0)
            new_pdf = PdfReader(packet)
            pdf_page.merge_page(new_pdf.pages[0])

        pdf_writer.add_page(pdf_page)

    packet_new_page = BytesIO()
    new_page_canvas = canvas.Canvas(packet_new_page, pagesize=A4)
    add_last_page_text(new_page_canvas)

    new_page_number_text = f"Page {total_pages + 1} of {total_pages + 1}"
    new_page_canvas.setFont("Montserrat-Regular", 8)
    new_page_canvas.setFillColor(colors.HexColor("#545454"))
    new_page_canvas.drawString(510, 10, new_page_number_text)

    new_page_canvas.save()
    packet_new_page.seek(0)
    new_pdf_page = PdfReader(packet_new_page)
    pdf_writer.add_page(new_pdf_page.pages[0])

    pdf_writer.write(buffer_2)

    buffer_2.seek(0)

    buffer.close()

    if hasattr(pi, "convertedpi"):
        if pi.convertedpi.is_taxInvoice:
            filename = f"Tax Invoice_{pi.company_name}_{pi.convertedpi.invoice_no}_{pi.convertedpi.invoice_date}.pdf"
        else:
            filename = f"PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.pdf"
    else:
        filename = f"PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.pdf"

    response = HttpResponse(buffer_2.getvalue(), content_type="application/pdf")

    response["Content-Disposition"] = f"attachment; filename={filename}"

    return response
