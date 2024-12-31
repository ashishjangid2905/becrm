import os
import re
from pathlib import Path

# all Django imports
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseForbidden
from django.urls import reverse
# from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db.models import Q, Min, Max
from django.contrib import messages
# from django.contrib.staticfiles import finders
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.mail import EmailMultiAlternatives

# All Import from apps
from .models import biller, bankDetail, proforma, orderList, convertedPI, processedOrder, pi_number, BillerVariable
from lead.models import leads, contactPerson
from teams.models import User, Profile, SmtpConfig, UserVariable
from sample.models import Portmaster, CountryMaster
from .utils import SUBSCRIPTION_MODE, STATUS_CHOICES, PAYMENT_TERM, COUNTRY_CHOICE, STATE_CHOICE, CATEGORY, REPORT_TYPE, PAYMENT_STATUS, REPORT_FORMAT, ORDER_STATUS, REPORTS, FORMAT
from teams.custom_email_backend import CustomEmailBackend
from .templatetags.custom_filters import total_order_value, total_lumpsums, filter_by_lumpsum, total_pi_value_inc_tax
from .decorators import proforma_approval_required
from teams.templatetags.teams_custom_filters import get_current_position

# Third-party imports
import fiscalyear
from datetime import datetime, timedelta
from num2words import num2words
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.platypus import Image as Im
from pypdf import PdfReader, PdfWriter
from io import BytesIO

# from docx import Document
# from docx.shared import Pt, RGBColor
# from docx.enum.text import WD_PARAGRAPH_ALIGNMENT


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
# from openpyxl.drawing.image import Image
# from openpyxl.cell.text import InlineFont
# from openpyxl.cell.rich_text import TextBlock, CellRichText
# from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
# Create your views here.



@login_required(login_url='app:login')
def biller_list(request):
    billers = biller.objects.all()

    context = {
        'billers': billers,
    }
    return render(request, 'invoices/biller-list.html', context)


def get_biller_variable(biller_dtl,variable_name):
    today = timezone.now().date()

    filters = {'from_date__lte': today, 'biller_id': biller_dtl, 'variable_name': variable_name }

    variable = BillerVariable.objects.filter(
        Q(to_date__gte=today) | Q(to_date__isnull = True),
        **filters).last()

    return variable.variable_value if variable else None

@login_required(login_url='app:login')
def biller_detail(request, biller_id):

    biller_dtl = biller.objects.get(pk=biller_id)
    banks = bankDetail.objects.filter(biller_id = biller_id)
    format_choice = FORMAT

    biller_variables = {
        "pi_tag": get_biller_variable(biller_dtl, "pi_tag"),
        "pi_format": get_biller_variable(biller_dtl, "pi_format"),
        "invoice_tag": get_biller_variable(biller_dtl, "invoice_tag"),
        "invoice_format": get_biller_variable(biller_dtl, "invoice_format"),
    }

    context = {
        'biller_dtl': biller_dtl,
        'banks': banks,
        'format_choice': format_choice,
        'biller_variables': biller_variables

    }

    return render(request, 'invoices/biller.html', context)

@login_required(login_url='app:login')
def set_format(request, biller_id):
    if request.user.role != 'admin':
        return redirect('app:dashboard')
    
    biller_instanse = get_object_or_404(biller, pk=biller_id)
    user_id = request.user.id

    target_url = reverse('invoice:biller_detail', args=[biller_instanse.id])

    biller_variables = {
        "pi_tag": get_biller_variable(biller_instanse, "pi_tag"),
        "pi_format": get_biller_variable(biller_instanse, "pi_format"),
        "invoice_tag": get_biller_variable(biller_instanse, "invoice_tag"),
        "invoice_format": get_biller_variable(biller_instanse, "invoice_format"),
    }

    if request.method == 'POST':
        pi_tag = request.POST.get('pi_tag')
        pi_format = request.POST.get('pi_format')
        invoice_tag = request.POST.get('invoice_tag')
        invoice_format = request.POST.get('invoice_format')
        from_date_str = request.POST.get('from_date')

        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponseRedirect(target_url)

        for key, currentValue in biller_variables.items():
            if currentValue:
                existing_Variables = BillerVariable.objects.filter(biller_id = biller_instanse, variable_name = key).last()
                if existing_Variables:
                    existing_Variables.to_date = from_date - timedelta(days=1)
                    existing_Variables.save()

                newValue = request.POST.get(key)
                if newValue:
                    BillerVariable.objects.create(biller_id = biller_instanse, variable_name = key, variable_value = newValue, from_date = from_date, inserted_by = user_id )

    return HttpResponseRedirect(target_url)

@login_required(login_url='app:login')
def add_biller(request):

    if request.user.role == 'admin' and request.method == 'POST':
        user = request.user.id
        biller_name = request.POST.get("biller_name")
        brand_name = request.POST.get("brand_name")
        biller_gstin = request.POST.get("gstin")
        biller_msme = request.POST.get("msme")
        biller_pan = request.POST.get("pan")
        reg_address1 = request.POST.get("address1")
        reg_address2 = request.POST.get("address2")
        reg_city = request.POST.get("city")
        reg_state = request.POST.get("state")
        reg_pincode = request.POST.get("pincode")
        reg_country = request.POST.get("country")
        corp_address1 = request.POST.get("corp_address1")
        corp_address2 = request.POST.get("corp_address2")
        corp_city = request.POST.get("corp_city")
        corp_state = request.POST.get("corp_state")
        corp_pincode = request.POST.get("corp_pincode")
        corp_country = request.POST.get("corp_country")

        try:
            newBiller = biller.objects.create(biller_name = biller_name, brand_name=brand_name, biller_gstin = biller_gstin, biller_msme = biller_msme, biller_pan = biller_pan, reg_address1 = reg_address1, reg_address2 = reg_address2, reg_city = reg_city, reg_state = reg_state, reg_pincode = reg_pincode, reg_country = reg_country, corp_address1 = corp_address1, corp_address2 = corp_address2, corp_city = corp_city, corp_state = corp_state, corp_pincode = corp_pincode, corp_country = corp_country )
            return redirect('invoice:biller_list')
        except IntegrityError:
            messages.error(request, 'Biller with this GSTIN already exists.')
            return render(request, 'invoices/add-biller.html')

    return render(request, 'invoices/add-biller.html')

@login_required(login_url='app:login')
def add_bank(request, biler_id):

    biller_obj = biller.objects.get(pk=biler_id)
    if request.user.role == 'admin' and request.method == 'POST':
        bnf_name = request.POST.get('beneficiary_name')
        is_upi = request.POST.get("is_upi") == "on"
        try:
            if is_upi:
                # Handle UPI-specific bank details
                upi_id = request.POST.get("upi_id")
                upi_no = request.POST.get("upi_no")
                bankDetail.objects.create(
                    biller_id=biller_obj,
                    bnf_name=bnf_name,
                    is_upi=is_upi,
                    upi_id=upi_id,
                    upi_no=upi_no
                )
            else:
                # Handle traditional bank details
                bank_name = request.POST.get('bank_name')
                branch_address = request.POST.get('branch_address')
                ac_no = request.POST.get('ac_no')
                ifsc = request.POST.get('ifsc_code')
                swift_code = request.POST.get('swift_code')
                bankDetail.objects.create(
                    biller_id=biller_obj,
                    bnf_name=bnf_name,
                    bank_name=bank_name,
                    branch_address=branch_address,
                    ac_no=ac_no,
                    ifsc=ifsc,
                    swift_code=swift_code
                )

            messages.success(request, f"Bank details added successfully. {is_upi}")
            return redirect(reverse('invoice:biller_detail', args=[biler_id]))

        except IntegrityError:
            messages.error(request, "An error occurred while saving the bank details. Please try again.")
            return redirect(request.path_info)


    context={
        'biller_dtl': biller_obj,
    }
    return render(request, 'invoices/add-bank.html', context)


@login_required(login_url='app:login')
def pi_list(request):

    today = timezone.now().date()

    date_range = proforma.objects.aggregate(min_date = Min('pi_date'), max_date = Max('pi_date'))

    if date_range['min_date'] and ['max_date']:
        min_year = date_range['min_date'].year
        max_year = date_range['max_date'].year
    else:
        min_year = today.year
        max_year = today.year

    fiscal_years = []
    for year in range(min_year, max_year + 1):
        fy_label = f"{year}-{year + 1}"
        fiscal_years.append(fy_label)

    selected_fy = request.GET.get('fy', None)

    if selected_fy:
        # Parse the selected fiscal year from the format '2023-2024'
        fy_start_year = int(selected_fy.split('-')[0])
        fy_end_year = fy_start_year + 1
        
        fy_start = datetime(fy_start_year, 4, 1).date()
        fy_end = datetime(fy_end_year, 3, 31).date()
    else:
        # Default to the current fiscal year
        if today.month >= 4:  # We're in the current fiscal year
            fy_start = datetime(today.year, 4, 1).date()
            fy_end = datetime(today.year + 1, 3, 31).date()
        else:  # We're in the previous fiscal year
            fy_start = datetime(today.year - 1, 4, 1).date()
            fy_end = datetime(today.year, 3, 31).date()

    profile_instance = get_object_or_404(Profile, user=request.user)
    user_branch = profile_instance.branch
    all_users = Profile.objects.filter(user__profile__branch=user_branch.id)

    current_position = get_current_position(profile_instance)
    user_role = request.user.role

    if current_position == 'Head' or user_role == 'admin':
        all_users = all_users
    elif current_position == 'VP':
        all_users = all_users.exclude(user__groups__name='Head')
    elif current_position == 'Sr. Executive':
        all_users = all_users.exclude(user__groups__name__in=['Head', 'VP', 'Sr. Executive']).union(all_users.filter(user=request.user))

    selected_user = request.GET.get("user", None)
    selected_Ap = request.GET.get("ap", None)
    selected_status = request.GET.get("status", None)

    # Filter option logics

    filters = {'pi_date__gte': fy_start, 'pi_date__lte': fy_end}
    if selected_user:
        filters['user_id'] = selected_user

    if selected_Ap:
        filters['is_Approved'] = selected_Ap

    if selected_status:
        filters['status'] = selected_status

    # Impliment the applied filters

    all_proforma = proforma.objects.filter(**filters).order_by( '-pi_date','-pi_no')

    status_choices = STATUS_CHOICES
    payment_status = PAYMENT_STATUS
    report_type = REPORT_TYPE
    report_format = REPORT_FORMAT

    all_user_ids = []
    for user in all_users:
        all_user_ids.append(user.user.id)

    if request.user.role == 'admin' or current_position == 'Head' or current_position == 'VP' or current_position == 'Sr. Executive':
        all_pi = all_proforma.filter(user_id__in = all_user_ids)
    else:
        all_pi = all_proforma.filter(user_id = request.user.id)


    query = request.GET.get('q')

    search_fields = [
        'company_name', 'gstin', 'state', 'country', 'requistioner','email_id', 'contact', 'status',
        'pi_no', 'orderlist__product', 'orderlist__report_type'
    ]

    search_objects = Q()

    if query:
        # Build the Q object for searching leads
        for field in search_fields:
            search_objects |= Q(**{f'{field}__icontains': query})

        piList = all_pi.filter(search_objects).distinct()
    else:
        piList = all_pi

    if piList:
        for pi in piList:
            pi.user_id = get_object_or_404(User,pk = pi.user_id)

            if pi.approved_by is not None:
                pi.approved_by = get_object_or_404(User, pk = pi.approved_by)

    pageSize = request.GET.get('pageSize', 20)

    pi_per_page = Paginator(piList, pageSize)
    page = request.GET.get('page')

    try:
        piList = pi_per_page.get_page(page)
    except PageNotAnInteger:
        piList = pi_per_page.get_page(1)
    except EmptyPage:
        piList = pi_per_page.get_page(pi_per_page.num_pages)

    context = {
        'all_users': all_users,
        'all_pi': piList,
        'status_choices': status_choices,
        'payment_status': payment_status,
        'report_type': report_type,
        'report_format': report_format,
        'fiscal_years': fiscal_years,
        'selected_user': selected_user,
        'selected_Ap': selected_Ap,
        'selected_status': selected_status,
        'selected_fy': selected_fy,
        'pageSize': pageSize,
        'all_user_ids': all_user_ids,
        'current_position': current_position,
        'user_role': user_role
    }
    return render(request, 'invoices/proforma-list.html', context)


@login_required(login_url='app:login')
def create_pi(request, lead_id=None):

    bank_choice = bankDetail.objects.all()
    subs_choice = SUBSCRIPTION_MODE
    pay_choice = PAYMENT_TERM
    status_choice = STATUS_CHOICES
    country_choice = COUNTRY_CHOICE
    state_choice = STATE_CHOICE
    category_choice = CATEGORY
    report_choice = REPORT_TYPE

    user_id = request.user.id

    all_leads = leads.objects.filter(user = user_id).order_by('company_name')

    lead_info = None
    contact_info = None
    target_url = reverse('invoice:pi_list')
    company_ref = None


    if request.method == 'GET':
        company_ref = request.GET.get('lead_ref')
        if company_ref:
            target_url = reverse('invoice:create_pi_lead_id', args=[company_ref])
            return redirect(target_url)
        elif company_ref == '':
            target_url = reverse('invoice:create_pi')
            return redirect(target_url)


    if lead_id:
        lead_info = get_object_or_404(leads, pk = lead_id)
        contact_info = lead_info.contactperson_set.filter(is_active=True).first()

        target_url = reverse('lead:leads_pi', args=[lead_id])

        if lead_info.user != request.user.id:
            return redirect('invoice:create_pi')

    if request.method == 'POST':
        user = request.user

        user_id = user.id
        company_ref = lead_info

        company_name = request.POST.get('company_name')
        gstin = request.POST.get('gstin')
        is_sez = request.POST.get('is_sez') == 'on'
        lut_no = request.POST.get('lut_no')
        vendor_code = request.POST.get('vendor_code')
        currency = request.POST.get('currency')
        po_no = request.POST.get('po_no')
        po_date_str = request.POST.get('po_date')
        subscription = request.POST.get('subs_mode')
        bank = request.POST.get('bank')
        bank_instance = get_object_or_404(bankDetail,pk = bank)
        payment_term = request.POST.get('payment_term')
        requistioner = request.POST.get('requistioner')
        email_id = request.POST.get('email')
        contact = request.POST.get('contact_no')
        country = request.POST.get('country')
        state = request.POST.get('state')
        address = request.POST.get('address')
        details = request.POST.get('details')

        po_date = None

        if po_date_str:
            try:
                po_date = datetime.strptime(po_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Invalid date format. It must be in YYYY-MM-DD format.')


        pi_tag = get_biller_variable(bank_instance.biller_id, "pi_tag")
        pi_format = get_biller_variable(bank_instance.biller_id, "pi_format")


        pi_no = pi_number(user, pi_tag, pi_format)

        clean_details = re.sub(r'</?(table|tbody|thead|tr|td|th)[^>]*>', ',', details)
        clean_details = re.sub(r',+', ', ', clean_details)

        newPI = proforma.objects.create(company_ref=company_ref, user_id=user_id,pi_no=pi_no, company_name=company_name, gstin=gstin, is_sez=is_sez, lut_no=lut_no,
                                        vendor_code=vendor_code, currency=currency, po_no=po_no, po_date=po_date, subscription=subscription,
                                        bank=bank_instance, payment_term=payment_term, requistioner=requistioner, email_id=email_id,
                                        contact=contact, address=address, country=country, state=state, details=clean_details)
        
        for i in range(0, len(request.POST.getlist('orders'))):
            category = request.POST.getlist(f'category')[i]
            report_type = request.POST.getlist(f'report_type')[i]
            product = request.POST.getlist('product')[i]
            from_month = request.POST.getlist(f'from_month')[i]
            to_month = request.POST.getlist(f'to_month')[i]
            unit_price = request.POST.getlist(f'unit_price')[i] or 0
            total_price = request.POST.getlist(f'total_price')[i] or 0
            is_lumpsum = request.POST.getlist(f'is_lumpsum')[i] == 'on' if i < len(request.POST.getlist('is_lumpsum')) else False
            lumpsum_amt = request.POST.getlist(f'lumpsum_amt')[i] or 0

            newOrder = orderList.objects.create(proforma_id = newPI, category=category, report_type=report_type, product=product,
                                            from_month=from_month, to_month=to_month, unit_price=int(unit_price), total_price=int(total_price),is_lumpsum=is_lumpsum, lumpsum_amt=int(lumpsum_amt)
                                            )

        return HttpResponseRedirect(target_url)
    if lead_info:
        lead_info.state = int(lead_info.state) if lead_info.state else None
    
    context = {
        'bank_choice': bank_choice,
        'subs_choice': subs_choice,
        'pay_choice': pay_choice,
        'status_choice': status_choice,
        'country_choice': country_choice,
        'state_choice': state_choice,
        'category_choice': category_choice,
        'report_choice': report_choice,
        'lead_info': lead_info,
        'company_ref': lead_id,
        'contact_info': contact_info,
        'all_leads': all_leads
    }

    return render(request, 'invoices/new-proforma.html', context)

@login_required
@proforma_approval_required
def approve_pi(request, pi_id):
    pi_instance = get_object_or_404(proforma, pk=pi_id)
    user_ = get_object_or_404(Profile, user=request.user)
    user_of_proforma = get_object_or_404(User, pk=pi_instance.user_id)

    if request.method == 'POST':
        is_Approved = request.POST.get('is_approved')
        is_approved = is_Approved == 'true'

        if request.user.role == 'admin' or user_.user.groups.filter(name='Head').exists():
            # Admin and Head can approve all proformas
            pi_instance.is_Approved = is_approved
            pi_instance.approved_by = request.user.id
            pi_instance.save()
            return redirect('invoice:pi_list')

        elif user_.user.groups.filter(name='VP').exists():
            # VP can approve all proformas except those owned by Head
            if user_of_proforma.groups.filter(name='Head').exists():
                return HttpResponseForbidden("You cannot approve Head's Proforma.")
            pi_instance.is_Approved = is_approved
            pi_instance.approved_by = request.user.id
            pi_instance.save()
            return redirect('invoice:pi_list')

        elif user_.user.groups.filter(name='Sr. Executive').exists():
            # Sr. Executive can approve all proformas except those owned by Head or VP
            if user_of_proforma.groups.filter(name='Head').exists() or user_of_proforma.groups.filter(name='VP').exists():
                return HttpResponseForbidden("You cannot approve Head's or VP's Proforma.")
            pi_instance.is_Approved = is_approved
            pi_instance.approved_by = request.user.id
            pi_instance.save()
            return redirect('invoice:pi_list')

    return redirect('invoice:pi_list')


@login_required(login_url='app:login')
def edit_pi(request, pi):

    bank_choice = bankDetail.objects.all()
    subs_choice = SUBSCRIPTION_MODE
    pay_choice = PAYMENT_TERM
    status_choice = STATUS_CHOICES
    country_choice = COUNTRY_CHOICE
    state_choice = STATE_CHOICE
    category_choice = CATEGORY
    report_choice = REPORT_TYPE

    pi_instance = get_object_or_404(proforma, slug=pi)
    existing_orders = pi_instance.orderlist_set.all()

    if pi_instance.company_ref:
        target_url = reverse('lead:leads_pi', args=[pi_instance.company_ref.id])
    else:
        target_url = reverse('invoice:pi_list' )

    if pi_instance.status == "closed":
        return redirect('invoice:pi_list')

    if pi_instance.user_id == request.user.id or request.user.role == 'admin':

        if request.method == 'POST':
            user = request.user

            user_id = user.id

            company_name = request.POST.get('company_name')
            gstin = request.POST.get('gstin')
            is_sez = request.POST.get('is_sez') == 'on'
            lut_no = request.POST.get('lut_no')
            currency = request.POST.get('currency')
            po_no = request.POST.get('po_no')
            po_date_str = request.POST.get('po_date')
            subscription = request.POST.get('subs_mode')
            bank = request.POST.get('bank')
            bank_instance = get_object_or_404(bankDetail,pk = bank)
            payment_term = request.POST.get('payment_term')
            requistioner = request.POST.get('requistioner')
            email_id = request.POST.get('email')
            contact = request.POST.get('contact_no')
            country = request.POST.get('country')
            state = request.POST.get('state')
            address = request.POST.get('address')
            details = request.POST.get('details')
            edited_by = request.user.id
            po_date = None

            if po_date_str:
                try:
                    po_date = datetime.strptime(po_date_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Invalid date format. It must be in YYYY-MM-DD format.')

            clean_details = re.sub(r'</?(table|tbody|thead|tr|td|th)[^>]*>', ',', details)
            clean_details = re.sub(r',+', ', ', clean_details)

            pi_instance.company_name = company_name
            pi_instance.gstin = gstin
            pi_instance.is_sez = is_sez
            pi_instance.currency = currency
            pi_instance.po_no = po_no
            pi_instance.po_date = po_date
            pi_instance.subscription = subscription
            pi_instance.bank = bank_instance
            pi_instance.payment_term = payment_term
            pi_instance.requistioner = requistioner
            pi_instance.email_id = email_id
            pi_instance.contact = contact
            pi_instance.country = country
            pi_instance.state = state
            pi_instance.address = address
            pi_instance.details = clean_details
            pi_instance.status = 'open'
            pi_instance.closed_at = None
            pi_instance.is_Approved = False
            pi_instance.approved_by = None
            pi_instance.edited_by = edited_by
            pi_instance.save()

            existing_orders = pi_instance.orderlist_set.all()

            totalOrders = request.POST.getlist('orders')
            categories = request.POST.getlist('category')
            report_types = request.POST.getlist('report_type')
            products = request.POST.getlist('product')
            from_months = request.POST.getlist('from_month')
            to_months = request.POST.getlist('to_month')
            unit_prices = request.POST.getlist('unit_price')
            total_prices = request.POST.getlist('total_price')
            lumpsum_amts = request.POST.getlist('lumpsum_amt')
            is_lumpsums = request.POST.getlist('is_lumpsum')

            for i, order in enumerate(existing_orders):

                if i < len(totalOrders):
                    order.category = categories[i]
                    order.report_type = report_types[i]
                    order.product = products[i]
                    order.from_month = from_months[i]
                    order.to_month = to_months[i]

                    order.unit_price = unit_prices[i]
                    order.total_price = total_prices[i]
                    order.is_lumpsum = is_lumpsums[i] == 'on' if i < len(request.POST.getlist('is_lumpsum')) else False
                    order.lumpsum_amt = lumpsum_amts[i]
                    order.save()
                else:
                    order.delete()

            for i in range(len(existing_orders), len(request.POST.getlist('orders'))):

                category = categories[i]
                report_type = report_types[i]
                product = products[i]
                from_month = from_months[i]
                to_month = to_months[i]
                unit_price = unit_prices[i]
                total_price = total_prices[i]
                is_lumpsum = is_lumpsums[i] == 'on' if i < len(request.POST.getlist('is_lumpsum')) else False
                lumpsum_amt = lumpsum_amts[i]

                newOrder = orderList.objects.create(proforma_id = pi_instance, category=category, report_type=report_type, product=product,
                                                from_month=from_month, to_month=to_month, unit_price=unit_price, total_price=total_price, is_lumpsum=is_lumpsum, lumpsum_amt=lumpsum_amt
                                                )

    
            if pi_instance.company_ref:
                return HttpResponseRedirect(target_url)
            else:
                return redirect('invoice:pi_list')
            
        context = {
            'bank_choice': bank_choice,
            'subs_choice': subs_choice,
            'pay_choice': pay_choice,
            'status_choice': status_choice,
            'country_choice': country_choice,
            'state_choice': state_choice,
            'category_choice': category_choice,
            'report_choice': report_choice,
            'pi_instance': pi_instance,
            'existing_orders': existing_orders,
            }
        return render(request, 'invoices/edit-proforma.html', context)


@login_required(login_url='app:login')
def update_pi_status(request, pi):
    
    pi_instance = get_object_or_404(proforma, pk=pi)

    target_url = reverse('invoice:pi_list' )

    if request.user.id == pi_instance.user_id and request.method == 'POST':
        pi_status = request.POST.get('pi_status')
        closed_at = request.POST.get('closingDate')
        pi_instance.closed_at = closed_at
        pi_instance.status = pi_status
        pi_instance.save()

        
        is_invoiceRequire = request.POST.get('is_invoiceRequire') == 'on'
        payment_status = request.POST.get('payment_status')
        payment1_date = request.POST.get('payment1_date')
        payment1_amt = request.POST.get('payment1_amt')
        is_closed = True

        closed_pi_instance = convertedPI.objects.filter(pi_id = pi_instance).first()
        if pi_instance.status == 'closed':
            if closed_pi_instance:
                try:
                    closed_pi_instance.is_invoiceRequire = is_invoiceRequire
                    closed_pi_instance.is_closed = is_closed
                    closed_pi_instance.is_hold = False
                    closed_pi_instance.is_cancel = False
                    closed_pi_instance.payment_status = payment_status
                    closed_pi_instance.payment1_date = payment1_date
                    closed_pi_instance.payment1_amt = payment1_amt
                    closed_pi_instance.save()
                except Exception as e:
                    messages.error(request, f'An error occurred: {str(e)}')
            else:
                convertedPI.objects.create(pi_id=pi_instance,is_invoiceRequire=is_invoiceRequire, 
                                           is_closed=is_closed, payment_status=payment_status,
                                           payment1_date=payment1_date, payment1_amt=payment1_amt)

        elif pi_instance.status == 'open' and closed_pi_instance:
            try:
                closed_pi_instance.is_hold = True
                closed_pi_instance.save()
            except Exception as e:
                messages.error(request, f'An error occurred: {str(e)}')
        elif pi_instance.status == 'lost' and closed_pi_instance:
            try:
                closed_pi_instance.is_invoiceRequire = False
                closed_pi_instance.is_hold = True
                closed_pi_instance.is_cancel = True
                closed_pi_instance.save()
            except Exception as e:
                messages.error(request, f'An error occurred: {str(e)}')

        return HttpResponseRedirect(target_url)
    messages.error(request, "You are not authorized user to Update Status")
    return HttpResponseRedirect(target_url)


@login_required(login_url='app:login')
def process_pi(request, pi):
    # Get the proforma instance and check if it's already closed
    pi_instance = get_object_or_404(proforma, pk=pi)
    closed_pi_instance = convertedPI.objects.filter(pi_id=pi_instance).first()
    port_choice = Portmaster.objects.all()
    country_choice = CountryMaster.objects.all()

    context = {
        'pi': pi_instance,
        'report_format':REPORT_FORMAT,
        'reports': REPORTS,
        'port_choice': port_choice,
        'country_choice': country_choice
    }

    if pi_instance.status == 'closed':

        if request.method == 'POST' and closed_pi_instance.is_processed == False:
            # Retrieve the lists of values from the form
            report_types = request.POST.getlist('report_type')
            report_formats = request.POST.getlist('report_format')
            countries = request.POST.getlist('country')
            hsns = request.POST.getlist('hsn')
            products = request.POST.getlist('product')
            iecs = request.POST.getlist('iec')
            exporters = request.POST.getlist('exporter')
            importers = request.POST.getlist('importer')
            from_months = request.POST.getlist('from_month')
            to_months = request.POST.getlist('to_month')

            foreign_countries_all = request.POST.getlist('foreign_country[]')
            ports_all = request.POST.getlist('ports[]')

            fc_index = 0
            ports_index = 0

            # Loop through each order entry, with index to handle multi-selects uniquely
            for i in range(len(report_types)):
                # Calculate the slice for current form entry
                foreign_countries = foreign_countries_all[fc_index:fc_index + len(report_types)]
                ports = ports_all[ports_index:ports_index + len(report_types)]

                processedOrder.objects.create(
                    pi_id=pi_instance,
                    report_type=report_types[i],
                    format=report_formats[i],
                    country=countries[i],
                    hsn=hsns[i],
                    product=products[i],
                    iec=iecs[i],
                    exporter=exporters[i],
                    importer=importers[i],
                    foreign_country=', '.join(foreign_countries),  # Join list into comma-separated string
                    port=', '.join(ports),  # Join list into comma-separated string
                    from_month=from_months[i],
                    to_month=to_months[i]
                )

                fc_index += len(report_types)
                ports_index += len(report_types)

            # Mark the converted PI instance as processed
            if closed_pi_instance:
                closed_pi_instance.is_processed = True
                closed_pi_instance.save()

            return redirect('invoice:processed_list')  # Redirect after processing

        return render(request, 'invoices/process-pi.html', context)
    
    return redirect('invoice:pi_list')


@login_required(login_url='app:login')
def processed_list(request):

    today = timezone.now().date()

    date_range = processedOrder.objects.aggregate(min_date = Min('order_date'), max_date = Max('order_date'))

    if date_range['min_date'] and ['max_date']:
        min_year = date_range['min_date'].year
        max_year = date_range['max_date'].year
    else:
        min_year = today.year
        max_year = today.year

    fiscal_years = []
    for year in range(min_year, max_year + 1):
        fy_label = f"{year}-{year + 1}"
        fiscal_years.append(fy_label)

    selected_fy = request.GET.get('fy', None)

    if selected_fy:
        # Parse the selected fiscal year from the format '2023-2024'
        fy_start_year = int(selected_fy.split('-')[0])
        fy_end_year = fy_start_year + 1
        
        fy_start = datetime(fy_start_year, 4, 1).date()
        fy_end = datetime(fy_end_year, 3, 31).date()
    else:
        # Default to the current fiscal year
        if today.month >= 4:  # We're in the current fiscal year
            fy_start = datetime(today.year, 4, 1).date()
            fy_end = datetime(today.year + 1, 3, 31).date()
        else:  # We're in the previous fiscal year
            fy_start = datetime(today.year - 1, 4, 1).date()
            fy_end = datetime(today.year, 3, 31).date()

    profile_instance = get_object_or_404(Profile, user=request.user)
    user_branch = profile_instance.branch
    all_users = Profile.objects.filter(user__profile__branch=user_branch.id, user__department="sales")

    processed_pi_list = proforma.objects.filter(convertedpi__is_processed = True)

    selected_user = request.GET.get("user", None)
    selected_Ap = request.GET.get("ap", None)
    selected_status = request.GET.get("status", None)

    filters = {'pi_date__gte': fy_start, 'pi_date__lte': fy_end}

    if selected_user:
        filters['user_id'] = selected_user

    if selected_Ap:
        filters['convertedpi__is_hold'] = selected_Ap
    
    if selected_status:
        filters['processedorder__order_status'] = selected_status

    all_processed_orders = processed_pi_list.filter(**filters).order_by('-convertedpi__requested_at')

    status_choices = ORDER_STATUS
    payment_status = PAYMENT_STATUS
    report_type = REPORT_TYPE
    report_format = REPORT_FORMAT

    all_user_ids = []
    for user in all_users:
        all_user_ids.append(user.user.id)

    if request.user.role == 'admin' or request.user.department == 'production':
        processed_orders = all_processed_orders.filter(user_id__in = all_user_ids)
    else:
        processed_orders = all_processed_orders.filter(user_id = request.user.id)



    query = request.GET.get('q')

    search_fields = [
        'company_name', 'gstin', 'requistioner','email_id', 'processedorder__report_type', 'processedorder__format',
        'pi_no', 'processedorder__hsn', 'processedorder__product', 'processedorder__port'
    ]

    search_objects = Q()

    if query:
        # Build the Q object for searching leads
        for field in search_fields:
            search_objects |= Q(**{f'{field}__icontains': query})

        processedOrderList = processed_orders.filter(search_objects).distinct()
    else:
        processedOrderList = processed_orders.distinct()

    if processedOrderList:
        for pi in processedOrderList:
            pi.user_id = get_object_or_404(User, pk=pi.user_id)

    pageSize = request.GET.get('pageSize', 20)

    pi_per_page = Paginator(processedOrderList, pageSize)
    page = request.GET.get('page')

    try:
        processedOrderList = pi_per_page.get_page(page)
    except PageNotAnInteger:
        processedOrderList = pi_per_page.get_page(1)
    except EmptyPage:
        processedOrderList = pi_per_page.get_page(pi_per_page.num_pages)

    context = {
        'processed_pi': processedOrderList,
        'all_users': all_users,
        'selected_user': selected_user,
        'selected_Ap':selected_Ap,
        'selected_status': selected_status,
        'selected_fy': selected_fy,
        'status_choices': status_choices,
        'fiscal_years': fiscal_years,
        'pageSize': pageSize,
    }

    return render(request, 'invoices/processed-orders.html', context)

@login_required(login_url='app:login')
def processed_pi_order(request, pi_id):
    pi_instance = get_object_or_404(proforma, pk=pi_id)

    context = {
        'pi': pi_instance,
        'order_status': ORDER_STATUS,
    }
    return render(request, 'invoices/processed-pi.html', context)

@login_required(login_url='app:login')
def update_order_status(request, order_id):
    order_instance = get_object_or_404(processedOrder, pk=order_id)

    if request.user.department == 'production' and request.method == 'POST':
        last_dispatch_month = request.POST.get('last_dispatch_month')
        last_dispatch_date = request.POST.get('last_dispatch_date')
        order_status = request.POST.get('order_status')
        
        if last_dispatch_month:
            order_instance.last_dispatch_month = last_dispatch_month

        if last_dispatch_date:
            last_dispatch_date = datetime.strptime(last_dispatch_date, "%Y-%m-%d").date()
            order_instance.last_dispatch_date = last_dispatch_date

        order_instance.order_status = order_status
        order_instance.last_sent_date = timezone.now().date()
        order_instance.save()

        messages.success(request, "Order Status Updated Successfully")
        return redirect(reverse('invoice:processed_pi_order', args=[order_instance.pi_id.id]))
    messages.error(request, "You are not autharized person to update Order Status")
    return redirect(reverse('invoice:processed_pi_order', args=[order_instance.pi_id.id]))


@login_required(login_url='app:login')
def invoice_list(request):

    today = timezone.now().date()

    date_range = convertedPI.objects.aggregate(min_date = Min('invoice_date'), max_date = Max('invoice_date'))

    if date_range['min_date'] and ['max_date']:
        min_year = date_range['min_date'].year
        max_year = date_range['max_date'].year
    else:
        min_year = today.year
        max_year = today.year

    fiscal_years = []
    for year in range(min_year, max_year + 1):
        fy_label = f"{year}-{year + 1}"
        fiscal_years.append(fy_label)

    selected_fy = request.GET.get('fy', None)

    if selected_fy:
        # Parse the selected fiscal year from the format '2023-2024'
        fy_start_year = int(selected_fy.split('-')[0])
        fy_end_year = fy_start_year + 1
        
        fy_start = datetime(fy_start_year, 4, 1).date()
        fy_end = datetime(fy_end_year, 3, 31).date()
    else:
        # Default to the current fiscal year
        if today.month >= 4:  # We're in the current fiscal year
            fy_start = datetime(today.year, 4, 1).date()
            fy_end = datetime(today.year + 1, 3, 31).date()
        else:  # We're in the previous fiscal year
            fy_start = datetime(today.year - 1, 4, 1).date()
            fy_end = datetime(today.year, 3, 31).date()

    profile_instance = get_object_or_404(Profile, user=request.user)
    user_branch = profile_instance.branch
    all_users = Profile.objects.filter(user__profile__branch=user_branch.id, user__department="sales")

    requestedInvoices = convertedPI.objects.filter(is_invoiceRequire = True)

    selected_user = request.GET.get("user", None)
    selected_Ap = request.GET.get("ap", None)
    selected_status = request.GET.get("status", None)

    filters = {'requested_at__gte': fy_start, 'requested_at__lte': fy_end}
    if selected_user:
        filters['pi_id__user_id'] = selected_user

    if selected_Ap:
        filters['is_invoiceRequire'] = selected_Ap
    
    if selected_status:
        filters['is_taxInvoice'] = selected_status

    all_invoices = requestedInvoices.filter(**filters).order_by('-requested_at')

    status_choices = ORDER_STATUS
    payment_status = PAYMENT_STATUS
    report_type = REPORT_TYPE
    report_format = REPORT_FORMAT

    all_user_ids = []
    for user in all_users:
        all_user_ids.append(user.user.id)

    if request.user.role == 'admin' or request.user.department == 'account':
        taxInvoices = all_invoices.filter(pi_id__user_id__in = all_user_ids)
    else:
        taxInvoices = all_invoices.filter(pi_id__user_id = request.user.id)
    
    query = request.GET.get('q')

    search_fields = [
        'pi_id__company_name', 'pi_id__gstin', 'pi_id__requistioner','pi_id__email_id', 'pi_id__contact',
        'pi_id__pi_no', 'pi_id__po_no', 'pi_id__vendor_code','invoice_no', 'irn'
    ]

    search_objects = Q()

    if query:
        # Build the Q object for searching leads
        for field in search_fields:
            search_objects |= Q(**{f'{field}__icontains': query})

        taxInvoicesList = taxInvoices.filter(search_objects)
    else:
        taxInvoicesList = taxInvoices

    if taxInvoicesList:
        for pi in taxInvoicesList:
            pi.pi_id.user_id = get_object_or_404(User, pk=pi.pi_id.user_id)

    if request.GET.get('export') == 'excel':
        return exportInvoicelist(taxInvoicesList)


    pageSize = request.GET.get('pageSize', 20)

    pi_per_page = Paginator(taxInvoicesList, pageSize)
    page = request.GET.get('page')

    try:
        taxInvoicesList = pi_per_page.get_page(page)
    except PageNotAnInteger:
        taxInvoicesList = pi_per_page.get_page(1)
    except EmptyPage:
        taxInvoicesList = pi_per_page.get_page(pi_per_page.num_pages)

    context = {
        'taxInvoicesList': taxInvoicesList,
        'all_users': all_users,
        'selected_user': selected_user,
        'selected_Ap':selected_Ap,
        'selected_status': selected_status,
        'selected_fy': selected_fy,
        'status_choices': status_choices,
        'fiscal_years': fiscal_years,
        'payment_status': payment_status,
    }

    return render(request, 'invoices/invoices-list.html', context)


def get_invoice_no(biller_id, invoice_tag, invoice_format, new_no=None):
    fiscalyear.START_MONTH = 4
    fy = str(fiscalyear.FiscalYear.current())[-2:]
    py = str(int(fy) - 1)
    
    if not isinstance(invoice_format, str):
        raise ValueError(f"Expected pi_format to be a string, got {type(invoice_format)}")

    search_prefix = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=0).rstrip("0")

    if not new_no:
        last_invoice = convertedPI.objects.filter(invoice_no__startswith=search_prefix, pi_id__bank__biller_id = biller_id).order_by('pi_no').last()

        if last_invoice:
            try:
                last_no = int(last_invoice.pi_no.split(search_prefix)[-1])
            except:
                last_no = 0
            new_no = last_no + 1
        else:
            new_no = 1

    invoice_no = invoice_format.format(py=py, fy=fy, tag=invoice_tag, num=new_no)

    return invoice_no



@login_required(login_url='app:login')
def generate_invoice(request, pi):
    if request.user.role != 'admin' and request.user.department != 'account':
        return redirect('invoice:invoice_list')
    
    invoice_instance = get_object_or_404(convertedPI, pk=pi)
    all_invoices = convertedPI.objects.filter(is_taxInvoice = True).order_by('-invoice_no')

    biller_id = invoice_instance.pi_id.bank.biller_id

    invoice_tag = get_biller_variable(biller_id, "invoice_tag")
    invoice_format = get_biller_variable(biller_id, "invoice_format")

    if not invoice_instance.is_invoiceRequire and invoice_instance.is_hold:
        return redirect('invoice:invoice_list')
    
    if not invoice_instance.pi_id.bank.biller_id.biller_gstin:
        messages.error(request, "Biller isn't GSTIN registerd, So can't generate Tax Invoice")
        return redirect('invoice:invoice_list')

    
    if request.method == 'POST':
        invoice_no = request.POST.get('invoice_no')
        invoice_date_str = request.POST.get('invoice_date')

        try:
            invoice_date = datetime.strptime(invoice_date_str, "%Y-%m-%d").date()
        except ValueError:
            return redirect('invoice:invoice_list')


        if all_invoices.filter(invoice_no=invoice_no).exists():
            messages.error(request, "Invoice No is already exist")
            return redirect('invoice:invoice_list')

        try:
            invoice_instance.invoice_no = get_invoice_no(biller_id, invoice_tag, invoice_format, int(invoice_no))
            invoice_instance.invoice_date = invoice_date
            invoice_instance.is_taxInvoice = True
            invoice_instance.save()
            messages.success(request, "Invoice generated Successfully")
        except FileExistsError:
            return HttpResponse("error")
    return redirect('invoice:invoice_list')



def exportInvoicelist(invoices):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoices'

    headers = [
        'S.N','Team Member','Company Name', 'GSTIN', 'Address', 'State', 'Country',
        'PI No', 'PI Date', 'Contact Person Name', 'Email', 'Contact Number', 'Bank', 'A/C No', 'Beneficry Name', 'Payment Status','1st Payment', '1st Payment Date','2nd Payment',
        '2nd Payment Date','3rd Payment', '3rd Payment Date', 'Amount',
        'Amount (inc. tax)', 'Total Received','Is Generated', 'Invoice No', 'Invoice Date'
    ]

    ws.append(headers)

    header_fill = PatternFill(fill_type='solid', start_color="0099CCFF", end_color="0099CCFF")
    header_font = Font(name='Calibri', size=11, bold=True, color="00000000")

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    s = 1

    for invoice in invoices:
        total_received = int(invoice.payment1_amt if invoice.payment1_amt != None else 0) + int(invoice.payment2_amt if invoice.payment2_amt != None else 0) + int(invoice.payment3_amt if invoice.payment3_amt != None else 0)

        row = [
            s ,invoice.pi_id.user_id.first_name +" "+ invoice.pi_id.user_id.last_name, invoice.pi_id.company_name, invoice.pi_id.gstin, invoice.pi_id.address, dict(STATE_CHOICE).get(int(invoice.pi_id.state)),
            invoice.pi_id.country, invoice.pi_id.pi_no, invoice.pi_id.pi_date, invoice.pi_id.requistioner, invoice.pi_id.email_id,
            invoice.pi_id.contact, invoice.pi_id.bank.bank_name, invoice.pi_id.bank.ac_no, invoice.pi_id.bank.bnf_name, 
            invoice.payment_status, invoice.payment1_amt, invoice.payment1_date, invoice.payment2_amt, invoice.payment2_date,
            invoice.payment3_amt, invoice.payment3_date, total_order_value(invoice.pi_id), total_pi_value_inc_tax(invoice.pi_id),
            total_received, invoice.is_taxInvoice, invoice.invoice_no, invoice.invoice_date
        ]

        ws.append(row)
        s+=1

    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length+2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    today = timezone.now().date()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=invoice-list{today}.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@login_required(login_url='app:login')
def bulkInvoiceUpdate(request):
    
    if request.user.role != 'admin' and request.user.department != 'account':
        messages.error(request, "You are not Authorized User to Update")
        return redirect('invoice:invoice_list')
    
    target_url = reverse("invoice:invoice_list")
    
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        if not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, "Invalid file type. Please upload an Excel (.xlsx) file.")
            return redirect('invoice:invoice_list')
        
        updated_count = 0
        skipped_count = 0
        skipped_rows = []

        try:
            wd = load_workbook(uploaded_file)
            ws = wd.active

            expected_headers = [
                'S.N','Team Member','Company Name', 'GSTIN', 'Address', 'State', 'Country',
                'PI No', 'PI Date', 'Contact Person Name', 'Email', 'Contact Number', 'Bank', 'A/C No', 'Beneficry Name', 'Payment Status','1st Payment', '1st Payment Date','2nd Payment',
                '2nd Payment Date','3rd Payment', '3rd Payment Date', 'Amount',
                'Amount (inc. tax)', 'Total Received','Is Generated', 'Invoice No', 'Invoice Date'
                ]
            
            headers = [cell.value for cell in ws[1]]

            if headers != expected_headers:
                messages.error(request, "Invalid file format. Please upload a file exported from the system.")
                return redirect('invoice:invoice_list')
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                pi_no = row[7]
                invoice_no = row[26]
                invoice_date = row[27]

                if not pi_no:
                    skipped_rows.append((*row, "Missing PI No"))
                    skipped_count += 1
                    continue

                try:
                    update_invoice = convertedPI.objects.filter(pi_id__pi_no=pi_no).first()

                    biller_id = update_invoice.pi_id.bank.biller_id

                    invoice_tag = get_biller_variable(biller_id, "invoice_tag")
                    invoice_format = get_biller_variable(biller_id, "invoice_format")

                    new_invoice_no = get_invoice_no(biller_id, invoice_tag, invoice_format, int(invoice_no))

                    if not update_invoice:
                        skipped_rows.append((*row, f'User does not request Tax invoice for PI No {pi_no}'))
                        skipped_count += 1
                        continue

                    if update_invoice.is_taxInvoice and update_invoice.invoice_no != new_invoice_no:
                        skipped_rows.append((*row, f"Invoice already generate against PI and Invoice No is {update_invoice.invoice_no}"))
                        skipped_count += 1
                        continue

                    if convertedPI.objects.filter(invoice_no=new_invoice_no).exclude(pi_id__pi_no = pi_no).exists():
                        skipped_rows.append((*row, f'Invoice No {invoice_no} already Exists'))
                        skipped_count += 1
                        continue

                    update_invoice.invoice_no = new_invoice_no
                    update_invoice.invoice_date = invoice_date
                    update_invoice.is_taxInvoice = True
                    update_invoice.save()
                    updated_count += 1
                
                except Exception as e:
                    skipped_rows.append((*row, f"Error: {e}"))
                    skipped_count += 1

            if skipped_rows:
                skipped_wd = Workbook()
                skipped_ws = skipped_wd.active

                skipped_ws.title = "Not Uploaded"

                headers = [cell.value for cell in ws[1]] + ["Error Reason"]
                skipped_ws.append(headers)

                for skipped_row in skipped_rows:
                    skipped_ws.append(skipped_row)

                today = timezone.now().date()
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=Not Uploaded-list{today}.xlsx'

                # Save the workbook to the response object
                skipped_wd.save(response)

                return response

            messages.success(request, f"Total Invoice Update {updated_count} and Error {skipped_count}")
            return HttpResponseRedirect(target_url)

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('invoice:invoice_list')
    messages.error(request, f"file is not Uploaded")
    return redirect('invoice:invoice_list')
    



def int_to_roman(num):
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4,
        1
    ]
    syms = [
        "m", "cm", "d", "cd",
        "c", "xc", "l", "xl",
        "x", "ix", "v", "iv",
        "i"
    ]
    
    roman_num = ""
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman_num += syms[i]
            num -= val[i]
        i += 1
    return roman_num



def pdf_PI(pi):
    pi = get_object_or_404(proforma, pk = pi)
    orders = orderList.objects.filter(proforma_id= pi)

    user_obj = get_object_or_404(User, pk=pi.user_id)  # Fetch the User object
    user = get_object_or_404(Profile, user=user_obj)

    reg_address = pi.bank.biller_id.get_reg_full_address()
    corp_address = pi.bank.biller_id.get_corp_full_address()

    net_total = total_order_value(pi)
    lumpsumAmt = total_lumpsums(pi)

    if pi.currency == 'inr':
        curr = '₹'
    else:
        curr = '$'

    if pi.is_sez:
        cgst = 0
        sgst = 0
        igst = 0
        total_inc_tax = net_total
    else:
        if pi.bank.biller_id.biller_gstin:
            if str(pi.state) == pi.bank.biller_id.biller_gstin[0:2]:
                cgst = net_total*0.09
                sgst = net_total*0.09
                igst = 0
                total_inc_tax = net_total*1.18
            else:
                cgst = 0
                sgst = 0
                igst = net_total*0.18
                total_inc_tax = net_total*1.18
        else:
            cgst = 0
            sgst = 0
            igst = 0
            total_inc_tax = net_total


    roundOff = total_inc_tax - round(total_inc_tax,0)

    if pi.currency == 'inr':
        total_val_words = f'Rs. {num2words(round(total_inc_tax,0), lang='en').replace(',', '').title()} Only'
    else:
        total_val_words = f'Usd {num2words(round(total_inc_tax,0), lang='en').replace(',', '').title()} Only'

    basedir = Path(__file__).resolve().parent.parent

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0*inch, rightMargin=0*inch, topMargin=0.2 * inch, bottomMargin=3 * inch,)

    pdfmetrics.registerFont(TTFont('Montserrat-Light', os.path.join(basedir,'static/becrm/fonts/Montserrat-Light.ttf')))
    pdfmetrics.registerFont(TTFont('Montserrat-Regular', os.path.join(basedir,'static/becrm/fonts/Montserrat-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('Montserrat-Medium', os.path.join(basedir,'static/becrm/fonts/Montserrat-Medium.ttf')))
    pdfmetrics.registerFont(TTFont('Montserrat-Bold', os.path.join(basedir,'static/becrm/fonts/Montserrat-Bold.ttf')))
    pdfmetrics.registerFont(TTFont('Quicksand-Bold', os.path.join(basedir,'static/becrm/fonts/Quicksand-Bold.ttf')))
    pdfmetrics.registerFont(TTFont('Quicksand-Light', os.path.join(basedir,'static/becrm/fonts/Quicksand-Light.ttf')))
    pdfmetrics.registerFont(TTFont('Quicksand-Medium', os.path.join(basedir,'static/becrm/fonts/Quicksand-Medium.ttf')))
    pdfmetrics.registerFont(TTFont('Quicksand-Regular', os.path.join(basedir,'static/becrm/fonts/Quicksand-VariableFont_wght.ttf')))
    pdfmetrics.registerFont(TTFont('Quicksand-SemiBold', os.path.join(basedir,'static/becrm/fonts/Quicksand-SemiBold.ttf')))
    pdfmetrics.registerFont(TTFont('Roboto-Medium', os.path.join(basedir,'static/becrm/fonts/Roboto-Medium.ttf')))
    pdfmetrics.registerFont(TTFont('Roboto-Regular', os.path.join(basedir,'static/becrm/fonts/Roboto-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('Roboto-Bold', os.path.join(basedir,'static/becrm/fonts/Roboto-Bold.ttf')))
    pdfmetrics.registerFont(TTFont('Roboto-Light', os.path.join(basedir,'static/becrm/fonts/Roboto-Light.ttf')))


    blue_font = ParagraphStyle(name="BlueStyle", fontSize=24, alignment=0, fontName='Helvetica-Bold', leading=20, underlineWidth=2 )
    black_font = ParagraphStyle(name="BlueStyle", fontSize=12, alignment=0, fontName='Montserrat-Light' )

    font_xxs = ParagraphStyle(name="font_xxs", fontSize=8, fontName='Quicksand-bold', alignment=0, leading=8, textColor=colors.HexColor("#ffffff"))
    font_xs = ParagraphStyle(name="font_xs", fontSize=7, fontName='Quicksand-Bold', alignment=0, leading=7, textColor=colors.HexColor("#ffffff"))
    font_s = ParagraphStyle(name="font_s", fontSize=10, fontName='Quicksand-light', alignment=0)
    font_m = ParagraphStyle(name="font_m", fontSize=11, fontName='Quicksand-light', alignment=0)
    font_l = ParagraphStyle(name="font_l", fontSize=12, fontName='Quicksand-light', alignment=0)

    brand_text = f'{pi.bank.biller_id.brand_name}'
    brand_name = Paragraph(f'<font color="#3182d9">{brand_text}</font>', blue_font)

    if pi.bank.biller_id.biller_name != pi.bank.biller_id.brand_name:
        founder_text = f'{pi.bank.biller_id.biller_name}'
        founder_name = Paragraph(f'<font color="#000000" >founded by {founder_text}</font>', black_font)
    else:
        founder_name = ""

    imagepath = os.path.join(basedir,'static/becrm/image/pi_back.png')
    logo = Im(imagepath, 8.25*inch,3*inch)

    taxAmount = Paragraph(f'<font>₹ {total_inc_tax}</font>', black_font)

    biller_font = ParagraphStyle(name="BlueStyle", fontSize=11, alignment=0, textColor='white', fontName='Roboto-Bold' )

    if corp_address:
        corpAddress = Paragraph(f'<strong>Corporate Office: </strong><font>{corp_address}</font>', font_xs)
    else:
        corpAddress = ""
    regAddress = Paragraph(f'<b>Reg. Office: </b><font>{reg_address}</font>', font_xs)
    gstin_text = f'<b>GSTIN: </b><font>{pi.bank.biller_id.biller_gstin}</font>' if corp_address else ''

    biller_Name = Paragraph(f"<b>{pi.bank.biller_id.biller_name} | {gstin_text}</b>", biller_font)

    user_dtl = None
    if pi.bank.biller_id.biller_gstin != '':
        user_dtl = Paragraph(f"<b>Email: </b><font>{user.user.email} || </font><b>Contact: </b><font>{user.phone}</font><br/><b>Website: </b><font>www.besmartexim.com</font>", biller_font)

    address = Paragraph(f"<b>Address: </b><font>{pi.address}</font>", font_xxs)

    email_symbol = "\u2709"
    email = Paragraph(f'<b>E: </b><font>{pi.email_id}</font>', font_xxs)
    contact = Paragraph(f"<b>M: </b><font>{pi.contact}</font>", font_xxs)

    piDate = pi.pi_date.strftime("%d-%b-%y")

    user_name = Paragraph(f"<font>{user.user.first_name} {user.user.last_name}</font>", font_m)
    requistioner = Paragraph(f"<font>{pi.requistioner}</font>", font_m)
    subs = Paragraph(f"<font>{pi.subscription}</font>", font_m)
    pay_m = Paragraph(f"<font>{pi.payment_term}</font>", font_m)

    unit_price_h = Paragraph(f"<b><font>Unit Price ({pi.currency.upper()})</font></b>", font_s)
    total_price_h = Paragraph(f"<b><font>Total ({pi.currency.upper()})</font></b>", font_s)
    if hasattr(pi, 'convertedpi'):
        invoice_type = 'TAX INVOICE' if pi.convertedpi.is_taxInvoice else 'PRO FORMA INVOICE'
    else:
        invoice_type = 'PRO FORMA INVOICE'

    data = [
        [brand_name,'','','','','','','','','',''],
        [founder_name,'','','','','','','','','', invoice_type],
        ['','','','','','','','','','',''],
        ['','','','','',logo,'','','','',''],
        ['Issued by:','','','','','','','',f'₹ {round(total_inc_tax,0):.2f}','','',],
        [biller_Name,'','','','','','','','','','',],
        [regAddress,'','',corpAddress,'','','','','Total Payable Amount','',''],
        ['','','','','','','','','','',''],
        ['Customer Details:','','','','','','','','VENDOR CODE:' if pi.vendor_code else '',pi.vendor_code,''],
        [pi.company_name.upper(),'','','','','','','','LUT NO:'if pi.lut_no else '',pi.lut_no,''],
        [f'GSTIN: {pi.gstin}','','','','','','','','PO DATE:' if pi.po_date else '','',''],
        [address,'','','','','','','','PO NUMBER:' if pi.po_no else '','',''],
        ['','','','','','','','','PI DATE:',piDate,''],
        [email,'','','','','','','','PI NUMBER:',pi.pi_no,''],
        [contact,'','','','','','','','Invoice Date:' if invoice_type=='TAX INVOICE' else '',pi.convertedpi.invoice_date if invoice_type=='TAX INVOICE' else '',''],
        ['','','','','','','','','Invoice No.:' if invoice_type=='TAX INVOICE' else '',pi.convertedpi.invoice_no if invoice_type=='TAX INVOICE' else '',''],
        ['','','','','','','','','','',''],
        ['','','','','','','','','','',''],
        ['','S.N','ITEM DESCRIPTION','','','','','RATE','',f'TOTAL ({curr})',''],
        ['','','','','','','','','','',''],
    ]


    row_heights = [0.4*inch, 0.3*inch] + 3*[0.1*inch] + 15*[0.2*inch]
    col_widths = 11*[0.7*inch]

    table = Table(data, col_widths, row_heights)

    style = TableStyle([
        ('SPAN',(0,0),(5,0)),
        ('FONT',(0,1),(6,1),'Montserrat-Regular', 13),
        ('SPAN',(0,1),(5,1)),
        ('VALIGN',(0,0),(-1,-1), 'TOP'),
        ('VALIGN',(7,1),(10,1),'BOTTOM'),
        ('TEXTCOLOR',(7,1),(10,1), colors.HexColor('#545454')),
        ('ALIGN',(0,0),(-1,-1), 'LEFT'),
        ('ALIGN',(5,3),(5,3), 'CENTER'),
        ('ALIGN',(7,1),(10,4), 'RIGHT'),
        ('FONT',(7,1),(10,3),'Montserrat-Regular', 22),
        ('FONT',(7,3),(10,4),'Roboto-Medium', 16),
        ('TEXTCOLOR',(7,3),(10,4),colors.HexColor('#545454')),
        ('VALIGN',(0,4),(6,4),'MIDDLE',),
        ('FONT',(0,4),(6,4),'Montserrat-Medium', 8),
        ('TEXTCOLOR',(0,3),(6,4),colors.HexColor('#ffffff')),
        ('SPAN',(8,4),(10,4)),
        ('SPAN',(8,6),(10,6)),
        ('ALIGN',(7,4),(10,6), 'CENTER'),
        ('VALIGN',(7,6),(10,6), 'TOP'),
        ('FONT',(7,6),(10,6),'Montserrat-Light', 9),
        ('ALIGN',(7,3),(9,3), 'CENTER'),
        ('SPAN',(0,5),(5,5)),
        ('SPAN',(0,6),(2,6)),
        ('SPAN',(3,6),(5,6)),
        ('FONT',(0,8),(10,8),'Montserrat-Regular', 9),
        ('FONT',(0,9),(4,9),'Roboto-Bold', 13),
        ('FONT',(0,10),(4,10),'Roboto-Regular', 10),
        ('SPAN',(0,11),(5,11)),
        ('SPAN',(0,13),(6,13)),
        ('SPAN',(0,14),(5,14)),
        ('ALIGN',(8,8),(8,15),'RIGHT'),
        ('FONT',(8,8),(8,15),'Quicksand-Bold', 9),
        ('FONT',(9,8),(9,15),'Quicksand-Medium', 9),
        ('LINEAFTER',(6,9),(6,13), 2,colors.white),
        ('TEXTCOLOR',(0,7),(10,15), colors.HexColor('#ffffff')),
        ('FONT',(0,18),(10,18),'Quicksand-Bold', 11),
        ('ALIGN',(7,18),(9,19),'RIGHT'),
    ])

    table.setStyle(style)


    nr = 0
    s = 1
    order_list = []
    lumpsumRow = 0
    
    net_total = total_order_value(pi)
    lumpsumAmt = total_lumpsums(pi)

    for order in filter_by_lumpsum(orders,True):
        if order.is_lumpsum:
            cat = dict(CATEGORY).get(str(order.category))
            report = dict(REPORT_TYPE).get(str(order.report_type))
            product = order.product
            from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
            to_month = datetime.strptime(order.to_month, '%Y-%m').strftime("%b'%y")
            period = f'{from_month} - {to_month}'

            if pi.currency == 'inr':
                unitPrice = f'Lumpsum'
                totalPrice = f'{lumpsumAmt:.2f}'
            else:
                unitPrice = f'Lumpsum'
                totalPrice = f'{lumpsumAmt:.2f}'

            sac = f'SAC: 998399'

        from_m = datetime.strptime(order.from_month, '%Y-%m')
        to_m = datetime.strptime(order.to_month, '%Y-%m')

        orderDes = Paragraph(f'<font>{report} | {product} | Period: {from_month} - {to_month}</font>', font_s)
        no_months = (to_m.year - from_m.year)*12 + to_m.month - from_m.month + 1

        total_m = f'Total: {no_months}'

        order_list.extend([
            [f'{int_to_roman(s)})',f'{cat} | SAC: 998399','','','','','','',''],
            ['',orderDes,'','','','','','',''],
            ])
        nr+= 2
        s+= 1
        lumpsumRow+=2

    for order in filter_by_lumpsum(orders,False):
            cat = dict(CATEGORY).get(str(order.category))
            report = dict(REPORT_TYPE).get(str(order.report_type))
            product = order.product
            from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
            to_month = datetime.strptime(order.to_month, '%Y-%m').strftime("%b'%y")

            orderDes = Paragraph(f'<font>{report} | {product} | Period: {from_month} - {to_month}</font>', font_s)
            period = f'{from_month} - {to_month}'


            unitPrice = f'{curr} {order.unit_price}'
            totalPrice = f'{order.total_price:.2f}'

            sac = f'SAC: 998399'

            from_m = datetime.strptime(order.from_month, '%Y-%m')
            to_m = datetime.strptime(order.to_month, '%Y-%m')

            no_months = (to_m.year - from_m.year)*12 + to_m.month - from_m.month + 1

            total_m = f'Total: {no_months}'

            order_list.extend([
                [f'{int_to_roman(s)})',f'{cat} | SAC: 998399','','','','',unitPrice,'',totalPrice],
                ['',orderDes,'','','','','','',''],
                ])
            nr+= 2
            s+= 1    

    orderStyle = []

    if lumpsumRow>0:
        for r in range(0, lumpsumRow, 2):
            end_row = min(lumpsumRow - 1, r + 1)
            order_list[0][6] = "Lumpsum"
            order_list[0][8] = f'{lumpsumAmt:.2f}'
            for i in range(r + 1, end_row + 1):
                order_list[i][6] = ''
                order_list[i][8] = ''
            orderStyle.extend([
                ('SPAN',(6,r),(6, lumpsumRow-1)),
                ('SPAN',(8,r),(8, lumpsumRow-1)),
                ])
            
    if nr>lumpsumRow:
        for r in range(lumpsumRow,nr, 2):
            orderStyle.extend([
                ('SPAN',(6,r),(6, r+1)),
                ('SPAN',(8,r),(8, r+1)),
                ])

    for r in range(0, nr, 2):
        orderStyle.extend([
            ('LINEABOVE',(0,0),(-1,0),2,colors.HexColor('#c63545')),
            ('SPAN',(1,r),(5,r)),
            ('SPAN',(1,r+1),(5,r+1)),
            ('FONT',(1,r),(5,r),'Quicksand-Bold',10),
            ('FONT',(1,r+1),(5,r+1),'Quicksand-Medium',10),
            ('FONT',(6,r),(8, -1),'Quicksand-Medium',10),
            ('ALIGN',(6,r),(8, -1),'RIGHT'),
            ('RIGHTPADDING',(6,r),(6, -1),-10),
            ])

    orderStyle.append(('ROWBACKGROUNDS', (0, 0), (5, -1), [colors.white, colors.white, colors.HexColor('#f1f1f1'), colors.HexColor('#f1f1f1')]))
    
    order_table = Table(order_list, col_widths)

    order_table.setStyle(orderStyle)

    pdf_canvas = canvas.Canvas(buffer, pagesize=A4)

    
    # Draw something on the canvas (example)
    def add_canvas(canvas, doc):

        canvas.setLineWidth(1)  # Set the border width

        # You can also reset the font color if needed for further text
        canvas.setFillColor(colors.HexColor('#004aad'))  # Reset to black or any other color after drawing
        canvas.setStrokeColor(colors.white)  # Set stroke color for the border
        canvas.setLineWidth(1)
        canvas.roundRect(360, 165, 200, 95, 10, fill=1)

        canvas.setFillColor(colors.HexColor('#3182d9'))  # Reset to black or any other color after drawing
        canvas.rect(30,101, 150, 30, fill=1)
        canvas.rect(180,101, 160, 30, fill=1)
        canvas.rect(340,101, 120, 30, fill=1)
        canvas.rect(460,101, 100, 30, fill=1)

        canvas.setFillColor(colors.HexColor('#38b6ff'))  # Reset to black or any other color after drawing
        canvas.rect(30,70, 150, 30, fill=1)
        canvas.rect(180,70, 160, 30, fill=1)
        canvas.rect(340,70, 120, 30, fill=1)
        canvas.rect(460,70, 100, 30, fill=1)

        canvas.rect(30,39, 530, 30, fill=1)



    def add_text(canvas,doc):
        canvas.setFont("Quicksand-Medium", 10)
        canvas.setFillColor(colors.HexColor('#ffffff'))  # Set the font color to a shade of blue

        # Set font color for the tax strings
        canvas.drawRightString(440, 245, f"AMOUNT:")
        canvas.drawRightString(530, 245, f"{curr} {net_total:.2f}")
        canvas.drawRightString(440, 230, "IGST (18%):")
        canvas.drawRightString(530, 230, f"{curr} {igst:.2f}")
        canvas.drawRightString(440, 215, "CGST (9%):")  # Use different y-coordinates
        canvas.drawRightString(530, 215, f"{curr} {cgst:.2f}")  # Use different y-coordinates
        canvas.drawRightString(440, 200, "SGST (9%):")  # Use different y-coordinates
        canvas.drawRightString(530, 200, f"{curr} {sgst:.2f}")  # Use different y-coordinates


        canvas.setFont("Quicksand-Medium", 8)
        canvas.drawRightString(440, 188, "rounded off (+/-):")  # Use different y-coordinates
        canvas.drawRightString(530, 188, f"{curr} {roundOff:.2f}")  # Use different y-coordinates
        canvas.setFont("Quicksand-Bold", 12)
        canvas.drawRightString(440, 173, f"TOTAL:")    # Use different y-coordinates
        canvas.drawRightString(530, 173, f"{curr} {round(total_inc_tax,0):.2f}")   # Use different y-coordinates
        
        canvas.setFont("Quicksand-Regular", 11)
        canvas.setFillColor(colors.HexColor('#004aad'))  # Set the font color to a shade of blue
        canvas.drawRightString(350, 230, f"AMOUNT PAYABLE IN WORDS:")   # Use different y-coordinates

        words = total_val_words.split(" ")
        line = []
        current_line = ""

        for word in words:
            test_line = current_line + word + " "

            if canvas.stringWidth(test_line,'Quicksand-Bold', 12) <= 250:
                current_line = test_line
            else:
                line.append(current_line.strip())
                current_line = word + " "
        if current_line:
            line.append(current_line.strip())

        line_height = 13

        canvas.setFont("Quicksand-Bold", 12)
        for i, line in enumerate(line):

            canvas.drawRightString(350, 215 - (i * line_height), f"{line}")   # Use different y-coordinates
        
        canvas.setFont("Montserrat-Medium", 10)
        canvas.setFillColor(colors.HexColor('#3182d9'))  # Set the font color to a shade of blue
        canvas.drawRightString(540, 145, "Whether Tax is payable under REVERSE CHARGE: ")   # Use different y-coordinates
        canvas.setFont("Montserrat-Bold", 10)
        canvas.drawRightString(560, 145, "NO")   # Use different y-coordinates
        
        canvas.setFillColor(colors.HexColor('#ffffff'))  # Set the font color to a shade of blue
        canvas.setFont("Montserrat-Bold", 10)
        canvas.drawCentredString(105, 112, "Team Member")
        canvas.drawCentredString(260, 112, "Requisitioner")
        canvas.drawCentredString(400, 112, "Subscription Mode")
        canvas.drawCentredString(510, 112, "Payment Term")

        canvas.setFont("Montserrat-Medium", 9)
        canvas.drawCentredString(105, 80, f"{user}")
        canvas.drawCentredString(260, 80, f"{pi.requistioner}")
        canvas.drawCentredString(400, 80, f"{pi.subscription.capitalize()}")
        canvas.drawCentredString(510, 80, f"{pi.payment_term.capitalize()}")

        email_H_width = canvas.stringWidth('Email: ','Montserrat-Medium',9)
        email_width = canvas.stringWidth(f'{user.user.email}','Montserrat-Bold',9)
        contact_H_width = canvas.stringWidth(' | Contact: ','Montserrat-Medium',9)
        contact_width = canvas.stringWidth(f'{user.phone}','Montserrat-Bold',9)
        web_H_width = canvas.stringWidth(' | Website: ','Montserrat-Medium',9)
        web_width = canvas.stringWidth('www.besmartexim.com','Montserrat-Bold',9)

        start_point = 295 - (email_H_width + email_width + contact_H_width + contact_width + web_H_width + web_width) / 2

        
        canvas.setFont("Montserrat-Medium", 9)
        canvas.drawString(start_point, 50, 'Email: ')
        canvas.drawString(start_point + email_H_width + email_width, 50, ' | Contact: ')
        canvas.drawString(start_point + email_H_width + email_width + contact_H_width + contact_width, 50, ' | Website: ')
        canvas.setFont("Montserrat-Bold", 9)
        canvas.drawString(start_point + email_H_width, 50, f'{user.user.email}')
        canvas.drawString(start_point + email_H_width + email_width + contact_H_width, 50, f'{user.phone}')
        canvas.drawString(start_point + email_H_width + email_width + contact_H_width + contact_width + web_H_width, 50, 'www.besmartexim.com')
        
        canvas.setFont("Montserrat-Regular", 7)
        canvas.setFillColor(colors.HexColor('#545454'))
        canvas.drawCentredString(300, 10, f'This is Computer generated {invoice_type}')

    def add_last_page_text(canvas):

        canvas.setFillColor(colors.HexColor('#004aad'))  # Reset to black or any other color after drawing
        canvas.setStrokeColor(colors.white)  # Set stroke color for the border
        canvas.setLineWidth(1)
        canvas.roundRect(30, 640, 290, 170, 10, fill=1)

        canvas.setFont("Montserrat-Medium", 11)
        canvas.setFillColor(colors.HexColor('#ffffff'))
        canvas.drawString(50, 780, "Kindly pay in favor of")
        if pi.bank.is_upi:
            canvas.drawRightString(145, 740, "Holder Name: ")
            canvas.drawRightString(145, 725, "UPI ID.: ")
            canvas.drawRightString(145, 710, "UPI No.: ")
        else:
            canvas.drawRightString(145, 740, "Bank Name: ")
            canvas.drawRightString(145, 725, "A/C No.: ")
            canvas.drawRightString(145, 710, "Branch Address: ")
            canvas.drawRightString(145, 660, "IFSC: ")
            if pi.country != 'IN':
                canvas.drawRightString(145, 645, "SWIFT: ")
            canvas.setFont("Montserrat-Bold", 11)


        if pi.bank.is_upi:
            canvas.drawString(150, 740, f'{pi.bank.bnf_name}')
            canvas.drawString(150, 725, f'{pi.bank.upi_id}')
            canvas.drawString(150, 710, f'{pi.bank.upi_no}')
        else:
            canvas.drawString(50, 760, f'{pi.bank.bnf_name}')
            canvas.drawString(150, 740, f'{pi.bank.bank_name}')
            canvas.drawString(150, 725, f'{pi.bank.ac_no}')

            canvas.drawString(150, 660, f'{pi.bank.ifsc}')
            if pi.country != 'IN':
                canvas.drawString(150, 645, f'{pi.bank.swift_code}')

        def wrap_string(string, font_size,wrap_length):

            # Wraping Branch Address
            string = string.split(" ")
            line = []
            current_line = ""

            for word in string:
                test_line = current_line + word + " "
                if canvas.stringWidth(test_line,'Montserrat-Bold', font_size) <= wrap_length:
                    current_line = test_line
                else:
                    line.append(current_line.strip())
                    current_line = word + " "
            if current_line:
                line.append(current_line.strip())

            return line

        if not pi.bank.is_upi:
            line = wrap_string(pi.bank.branch_address,11,150)
            line_height = 13
            for i, line in enumerate(line):
                canvas.drawString(150, 710 - (i * line_height), f"{line}")

        canvas.setFont("Montserrat-Bold", 11)
        canvas.setFillColor(colors.HexColor('#545454'))
        canvas.drawString(30, 580, f'Other {pi.bank.biller_id.biller_name} Details for future reference')
        canvas.setFillColor(colors.HexColor('#3182d9'))
        canvas.drawString(30, 500, f'Terms & Conditions')

        canvas.setFont("Montserrat-Regular", 11)
        canvas.setFillColor(colors.HexColor('#545454'))
        canvas.circle(50, 553, 2, stroke=1, fill=1)
        canvas.drawString(55, 550, f'PAN: {pi.bank.biller_id.biller_pan}')
        canvas.circle(50, 533, 2, stroke=1, fill=1)
        canvas.drawString(55, 530, f'Udyam Registration Number: {pi.bank.biller_id.biller_msme if pi.bank.biller_id.biller_msme else 'N/A' }')

        canvas.setFont("Montserrat-Regular", 9)
        canvas.setFillColor(colors.HexColor('#3182d9'))
        canvas.circle(50, 483, 2, stroke=1, fill=1)
        canvas.drawString(55, 480, 'Payment has to be made 100% in advance')
        canvas.circle(50, 463, 2, stroke=1, fill=1)
        canvas.drawString(55, 460, f'Company Legal Name is {pi.bank.biller_id.biller_name}')
        canvas.circle(50, 443, 2, stroke=1, fill=1)
        canvas.drawString(55, 440, 'Any changes in the order will attract cost once order is sent based on PI approval by the client')
        canvas.circle(50, 423, 2, stroke=1, fill=1)
        canvas.drawString(55, 420, f'18% GST will be applicable for all transactions')
        canvas.circle(50, 403, 2, stroke=1, fill=1)
        canvas.drawString(55, 400, f'E-Invoice is applicable w.e.f. 1st August 2023 and hence once filed, no changes cannot be made after 24 hrs.')



    elements = []
    elements.append(table)
    elements.append(order_table)

    
    buffer.seek(0)

    doc.build(elements)

    original_pdf = PdfReader(buffer)
    total_pages = len(original_pdf.pages)

    buffer_2 = BytesIO()

    pdf_writer = PdfWriter()

    for page_num in range(total_pages):
        pdf_page = original_pdf.pages[page_num]
        packet1 = BytesIO()
        page_canvas = canvas.Canvas(packet1,pagesize=A4)
        page_number_text = f"Page {page_num + 1} of {total_pages+1}"
        page_canvas.setFont("Montserrat-Regular", 8)
        page_canvas.setFillColor(colors.HexColor("#545454"))
        page_canvas.drawString(510, 10, page_number_text)
        page_canvas.save()
        packet1.seek(0)
        page_with_number = PdfReader(packet1)
        pdf_page.merge_page(page_with_number.pages[0])

        if page_num == total_pages - 1:
            packet = BytesIO()
            pdf_canvas1 = canvas.Canvas(packet, pagesize=A4)
            add_canvas(pdf_canvas1, pdf_page)
            add_text(pdf_canvas1, pdf_page)

            pdf_canvas1.save()

            packet.seek(0)
            new_pdf = PdfReader(packet)
            pdf_page.merge_page(new_pdf.pages[0])
        
        pdf_writer.add_page(pdf_page)

    packet_new_page = BytesIO()
    new_page_canvas = canvas.Canvas(packet_new_page, pagesize=A4)
    add_last_page_text(new_page_canvas)

    new_page_number_text = f"Page {total_pages + 1} of {total_pages + 1}"
    new_page_canvas.setFont("Montserrat-Regular", 8)
    new_page_canvas.setFillColor(colors.HexColor("#545454"))
    new_page_canvas.drawString(510, 10, new_page_number_text)

    new_page_canvas.save()
    packet_new_page.seek(0)
    new_pdf_page = PdfReader(packet_new_page)
    pdf_writer.add_page(new_pdf_page.pages[0])

    pdf_writer.write(buffer_2)

    buffer_2.seek(0)

    buffer.close()

    return buffer_2.getvalue()


@login_required(login_url='app:login')
def download_pdf2(request, pi_id):
    
    pi = get_object_or_404(proforma, pk = pi_id)
    
    if hasattr(pi, 'convertedpi'):
        if pi.convertedpi.is_taxInvoice:
            filename = f'Tax Invoice_{pi.company_name}_{pi.convertedpi.invoice_no}_{pi.convertedpi.invoice_date}.pdf'
        else:
            filename= f'PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.pdf'
    else:
        filename= f'PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.pdf'

    response = HttpResponse(pdf_PI(pi_id), content_type = 'application/pdf')

    response['Content-Disposition'] = f'attachment; filename={filename}' 

    return response


@login_required(login_url='app:login')
def email_form(request, pi):
    pi_instance = get_object_or_404(proforma, pk=pi)
    smtp_details = SmtpConfig.objects.filter(user=request.user).first()
    password = smtp_details.email_host_password
    context = {
        'pi_instance': pi_instance,
        'smtp_details': password
    }

    if pi_instance.user_id != request.user.id:
        return HttpResponse("You are not Authorised to Process the order.")
    
    return render(request, 'invoices/emailForm.html', context)

@login_required(login_url='app:login')
def send_test_mail(request, pi):
    pi_instance = get_object_or_404(proforma, pk=pi)

    smtp_details = SmtpConfig.objects.filter(user=request.user).first()

    if pi_instance.user_id != request.user.id:
        return HttpResponse("You are not Authorised to Process the order.")

    if request.method == 'GET':
        to_mail = request.GET.get('to')
        subject = request.GET.get('mailSubject')
        msg = request.GET.get('mailMessage')

    smtp_settings = {
        'host': smtp_details.smtp_server,
        'port': smtp_details.smtp_port,
        'username': smtp_details.user.email,
        'password': smtp_details.email_host_password,
        'use_tls': smtp_details.use_tls,
    }


    html_message = f'{msg}'
    recipient_list=[email.strip() for email in to_mail.split(',')]

    email_backend = CustomEmailBackend(**smtp_settings)

    email = EmailMultiAlternatives(
            subject=subject,
            body='plain_message',
            from_email='info@besmartexim.com',
            to=recipient_list,  # Replace with actual recipient
            connection= email_backend
        )

    email.attach_alternative(html_message, "text/html")

    fileValue = pdf_PI(pi_instance.id)

    fileName = f'PI_{pi_instance.company_name}_{pi_instance.pi_no}_{pi_instance.pi_date}.pdf'

    email.attach(fileName, fileValue, "application/pdf" )

    email.send()
    return HttpResponse(f'status: success, message: Email has been sent')


# def set_custom_margins(document, left, right, top, bottom):
#     # Accessing the document's section and modifying its page margin settings
#     section = document.sections[0]

#     section.page_height = Inches(11.69)  # A4 height in inches
#     section.page_width = Inches(8.27)    # A4 width in inches

#     # Convert inches to twips (1 inch = 1440 twips)
#     section.left_margin = left * 1440
#     section.right_margin = right * 1440
#     section.top_margin = top * 1440
#     section.bottom_margin = bottom * 1440

    

# def download_doc(request, pi_id):

#     pi = get_object_or_404(proforma, pk = pi_id)
#     orders = orderList.objects.filter(proforma_id= pi_id)

#     user_obj = get_object_or_404(User, pk=pi.user_id)  # Fetch the User object
#     user = get_object_or_404(Profile, user=user_obj)

#     reg_address = pi.bank.biller_id.get_reg_full_address()
#     corp_address = pi.bank.biller_id.get_corp_full_address()

#     net_total = total_order_value(pi)
#     lumpsumAmt = total_lumpsums(pi)

#     if pi.currency == 'inr':
#         curr = '₹'
#     else:
#         curr = '$'

#     if pi.is_sez:
#         cgst = 0
#         sgst = 0
#         igst = 0
#         total_inc_tax = net_total
#     else:
#         if str(pi.state) == pi.bank.biller_id.biller_gstin[0:2]:
#             cgst = net_total*0.09
#             sgst = net_total*0.09
#             igst = 0
#             total_inc_tax = net_total*1.18
#         else:
#             cgst = 0
#             sgst = 0
#             igst = net_total*0.18
#             total_inc_tax = net_total*1.18

#     roundOff = total_inc_tax - round(total_inc_tax,0)

#     if pi.currency == 'inr':
#         total_val_words = f'Rs. {num2words(round(total_inc_tax,0), lang='en').replace(',', '').title()} Only'
#     else:
#         total_val_words = f'Usd {num2words(round(total_inc_tax,0), lang='en').replace(',', '').title()} Only'
#     # Obtain the PDF file (ensure pdf_PI returns the correct file path)
#     pdf_file = pdf_PI(pi_id)  # e.g., 'path/to/your/file.pdf'
    
#     basedir = Path(__file__).resolve().parent.parent

#     # Path where the DOCX file will be saved temporarily
#     docx_file = os.path.join(basedir,'static/becrm/PI_Format.docx') # Adjust this path as needed

#     doc = Document(docx_file)

#     # BrandName = doc.add_paragraph(f'{pi.bank.biller_id.brand_name}')
#     # brandRun = BrandName.runs[0]
#     # brandRun.font.size = Pt(24)
#     # brandRun.font.bold = True
#     # brandRun.font.color.rgb = RGBColor(49, 130, 217)
#     # BrandName.paragraph_format.space_after = Pt(0)
#     # founderName = doc.add_paragraph(f'founded by {pi.bank.biller_id.biller_name}').runs[0]
#     # founderName.font.size = Pt(13)

#     table = doc.add_table(rows=20, cols=11)
#     table.autofit = True
#     table.style = 'Table Grid'

#     brandCell = table.cell(0,0)
#     brandCellP = brandCell.paragraphs[0]
#     brandCellPr = brandCellP.add_run(f'{pi.bank.biller_id.brand_name}')
#     brandCellPr.font.size = Pt(23)
#     brandCellPr.font.name = "Montserrat-Medium"
#     brandCellPr.font.bold = True
#     brandCellPr.font.color.rgb = RGBColor(49,130,217)
#     founderCell = table.cell(1,0)
#     founderCellP = founderCell.paragraphs[0]
#     founderCellPr = founderCellP.add_run(f'founded by {pi.bank.biller_id.biller_name}')
#     founderCellPr.font.size = Pt(13)
#     founderCellPr.font.name = "Montserrat-Regular"


#     PI_head = table.cell(0,7).paragraphs[0]
#     PI_headF = PI_head.paragraph_format
#     PI_headR = PI_head.add_run('Pro Forma Invoice')
#     PI_headR.font.size = Pt(21)
#     PI_headF.space_before = Pt(12)
#     PI_headR.font.name = "Montserrat-Medium"
#     PI_head.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
#     PI_headR.font.color.rgb = RGBColor(94,94,94)

#     table.cell(0, 0).merge(table.cell(0, 4))
#     table.cell(1, 0).merge(table.cell(1, 4))
#     table.cell(0, 7).merge(table.cell(1, 10))


#     billerCellH = table.cell(3,0).paragraphs[0]
#     billerCellHr = billerCellH.add_run('Issued by:')
#     billerCellHr.font.size = Pt(8)
#     billerCellHr.font.name = "Montserrat-Regular"
#     billerCellHr.font.color.rgb = RGBColor(255,255,255)

#     billerCell = table.cell(4,0)
#     billerName = f'{pi.bank.biller_id.biller_name} ,| GSTIN: {pi.bank.biller_id.biller_gstin}'

#     billerCellP = billerCell.paragraphs[0]
#     billerCellPr1 = billerCellP.add_run(billerName.split(',')[0])
#     billerCellPr1.font.size = Pt(12)
#     billerCellPr1.font.name = "Roboto-Medium"
#     billerCellPr1.font.bold = True
#     billerCellPr1.font.color.rgb = RGBColor(255,255,255)
#     billerCellPr2 = billerCellP.add_run(billerName.split(',')[1])
#     billerCellPr2.font.name = "Roboto-Medium"
#     billerCellPr2.font.bold = False
#     billerCellPr2.font.color.rgb = RGBColor(255,255,255)

#     table.cell(4, 0).merge(table.cell(4, 4))

#     regAddCell = table.cell(5,0)
#     corpAddCell = table.cell(5,3)
#     regAdd = f'Reg. Office:| {reg_address}'
#     corpAdd = f'Corporate Office:| {corp_address}'

#     regAddCellP = regAddCell.paragraphs[0]
#     regAddCellPr1 = regAddCellP.add_run(regAdd.split('|')[0])
#     regAddCellPr1.font.size = Pt(8)
#     regAddCellPr1.font.name = "Roboto-Medium"
#     regAddCellPr1.font.bold = True
#     regAddCellPr1.font.color.rgb = RGBColor(255,255,255)
#     regAddCellPr2 = regAddCellP.add_run(regAdd.split('|')[1])
#     regAddCellPr2.font.size = Pt(8)
#     regAddCellPr2.font.bold = False
#     regAddCellPr2.font.color.rgb = RGBColor(255,255,255)

#     corpAddCellP = corpAddCell.paragraphs[0]
#     corpAddCellPr1 = corpAddCellP.add_run(corpAdd.split('|')[0])
#     corpAddCellPr1.font.size = Pt(8)
#     corpAddCellPr1.font.bold = True
#     corpAddCellPr1.font.color.rgb = RGBColor(255,255,255)
#     corpAddCellPr2 = corpAddCellP.add_run(corpAdd.split('|')[1])
#     corpAddCellPr2.font.size = Pt(8)
#     corpAddCellPr2.font.bold = False
#     corpAddCellPr2.font.color.rgb = RGBColor(255,255,255)
#     table.cell(5, 0).merge(table.cell(5, 2))
#     table.cell(5, 3).merge(table.cell(5, 5))


#     topAmtCell = table.cell(3,7).paragraphs[0]
#     topAmtCellR = topAmtCell.add_run(f'₹ {round(total_inc_tax,0):.2f}')
#     topAmtCellR.font.size = Pt(15)
#     topAmtCellR.font.name = "Roboto"
#     topAmtCellR.font.bold = True
#     topAmtCellR.font.color.rgb = RGBColor(48,48,48)
#     topAmtCell.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER

#     topAmtHCell = table.cell(4,7).paragraphs[0]
#     topAmtHCellR = topAmtHCell.add_run('Total Payable Amount')
#     topAmtHCellR.font.size = Pt(10)
#     topAmtHCellR.font.name = "Montserrat-Light"
#     topAmtHCellR.font.bold = False
#     topAmtHCellR.font.color.rgb = RGBColor(48,48,48)
#     topAmtHCell.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER

#     table.cell(3,7).merge(table.cell(3,10))
#     table.cell(4,7).merge(table.cell(4,10))


#     doc_io = BytesIO()
#     doc.save(doc_io)
#     doc_io.seek(0)  # Go to the start of the BytesIO object

#     # Prepare the response to download the DOCX file
#     response = HttpResponse(doc_io.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
#     response['Content-Disposition'] = 'attachment; filename=proforma_invoice.docx'
    

#     return response





# @login_required(login_url='app:login')
# def download_xls(request, pi_id):
#     pi = get_object_or_404(proforma, pk = pi_id)
#     # user = User.objects.get(pk=pi.user_id)
#     orders = orderList.objects.filter(proforma_id= pi.id)
#     user_profile = Profile.objects.get(user=pi.user_id)
#     context = {
#         'pi': pi,
#         'user': user_profile,
#     }

#     if request.user.role != 'admin' and pi.user_id != request.user.id:
#         return HttpResponse("You do not have permission to access/download this PI", status=403)


#     net_total = 0
#     for order in orders:
#         net_total+=order.total_price

#     if pi.is_sez:
#         cgst = 0
#         sgst = 0
#         igst = 0
#         total_inc_tax = net_total
#     else:
#         cgst = net_total*0.09
#         sgst = net_total*0.09
#         igst = net_total*0.18
#         total_inc_tax = net_total*1.18

#     total_val_words = num2words(total_inc_tax, lang='en').replace(',', '').title()

#     wb = Workbook()
#     ws = wb.active
#     ws.title = f'{pi.pi_no}'[6:]

#     basedir = Path(__file__).resolve().parent.parent
#     image_path = os.path.join(basedir,'static/becrm/image/PI_LOGO.png')
#     img = Image(image_path)

#     img.width = 90
#     img.height = 90

#     ws.add_image(img, 'A3')
#     ws['A3'].alignment = Alignment(indent=0.5)

#     ws.merge_cells('F1:J1')
#     header = ws['F1']
#     header.value = "PRO FORMA INVOICE"
#     header.font = Font(name='Microsoft YaHei UI', size=18, bold=True, color="00808000")
#     header.alignment = Alignment(horizontal="right", vertical="center")

#     strip1 = ws.merge_cells('A2:C2')
#     strip2 = ws.merge_cells('D2:J2')

#     ws.row_dimensions[2].height = 6
#     ws.row_dimensions[7].height = 6
#     ws.row_dimensions[8].height = 75
#     ws.row_dimensions[21].height = 6
#     ws.column_dimensions['B'].width = 10
#     ws.column_dimensions['J'].width = 10

#     ws['A2'].fill = PatternFill(fill_type='solid', start_color="0000CCFF", end_color="0000CCFF")
#     ws['D2'].fill = PatternFill(fill_type='solid', start_color="00FFCC00", end_color="00FFCC00")

#     ws.merge_cells('D4:J5')
#     brand_name = ws['D4']

#     blue = InlineFont(b=True, sz=26, u='single', color='0000CCFF')
#     orange = InlineFont(b=True, sz=26, u='single', color='00FFCC00')
#     brand_name.value = CellRichText([TextBlock(blue, 'BE '), TextBlock(orange, 'SMART EXIM')])
#     brand_name.alignment = Alignment(horizontal="right", vertical="bottom")

#     ws.merge_cells('G6:J6')
#     legal_name = ws['G6']
#     legal_name.value = 'founded by SUN TRADE INC'
#     legal_name.font = Font(name='Microsoft New Tai Lue', size=9, bold=True, color="00000080")
#     legal_name.alignment = Alignment(horizontal="right", vertical="bottom")

#     ws.merge_cells('A8:E8')
#     ws.merge_cells('F8:J8')

#     biller_details = ws['A8']
#     bold_i = InlineFont(b=True, i=True, sz=8)
#     normal_i = InlineFont(b=False, i=True, sz=8)
#     bold_i_s = InlineFont(b=True, i=True, sz=9)
#     normal_i_s = InlineFont(b=False, i=True, sz=9)
#     biller_details.value = CellRichText([TextBlock(bold_i,f'Issued by:\n{pi.bank.biller_id.biller_name}\n'),TextBlock(bold_i,'Reg. Office:'), TextBlock(normal_i,f'{pi.bank.biller_id.reg_address1}, {pi.bank.biller_id.reg_address2}, {pi.bank.biller_id.reg_city}- {pi.bank.biller_id.reg_pincode}, {pi.bank.biller_id.reg_state}, {pi.bank.biller_id.reg_country}\n'), TextBlock(bold_i,'Corporate Office:'), TextBlock(normal_i, f'{pi.bank.biller_id.corp_address1} {pi.bank.biller_id.corp_address2}, {pi.bank.biller_id.corp_city}- {pi.bank.biller_id.corp_pincode}, {pi.bank.biller_id.corp_state}, {pi.bank.biller_id.corp_country}\n'),TextBlock(bold_i,'GSTIN:'), TextBlock(normal_i, f'{pi.bank.biller_id.biller_gstin}')])
#     biller_details.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

#     user_details = ws['F8']
#     user_details.value = CellRichText([TextBlock(bold_i,'Email: '), TextBlock(normal_i,f'{user_profile.user.email} || '), TextBlock(bold_i,'Contact: '), TextBlock(normal_i,f'{user_profile.phone}\n'), TextBlock(bold_i,'Website: '), TextBlock(normal_i,'www.besmartexim.com')])
#     user_details.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
#     user_details.border = Border(left=Side(border_style='thin'))

#     ws['A10'].value = 'Customer Details:'
#     ws['A10'].font = Font(name='Calibri', size=10, bold=True, underline='single')

#     client_name = ws['A11']
#     client_name.value = f'{pi.company_name}'
#     client_name.font = Font(name='Calibri', size=12, bold=True)

#     client_gstin = ws['A12']
#     client_gstin.value = CellRichText([TextBlock(InlineFont(b=True, sz=9), 'GSTIN: '), TextBlock(InlineFont(sz=9), f'{pi.gstin}')])

#     ws.merge_cells('A13:E15')
#     address = ws['A13']
#     address.value = CellRichText([TextBlock(InlineFont(b=True, i=True, sz=9),'Address: '), TextBlock(InlineFont(b=False, i=True, sz=9), f'{pi.address}')])
#     address.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

#     email = ws['A16']
#     email.value = CellRichText([TextBlock(InlineFont(b=True, i=True, sz=9), 'Email: '), TextBlock(InlineFont(i=True, sz=9), f'{pi.email_id}')])

#     contact = ws['A17']
#     contact.value = CellRichText([TextBlock(InlineFont(b=True, i=True, sz=9), 'Contact: '), TextBlock(InlineFont(i=True, sz=9), f'{pi.contact}')])

#     ws['H11'].value = 'PI DATE:'
#     ws['H12'].value = 'PI NUMBER.:'
#     ws['H11'].font = Font(name='Calibri', size=8, bold=True)
#     ws['H12'].font = Font(name='Calibri', size=8, bold=True)
#     ws['H11'].alignment = Alignment(horizontal='right', vertical='center')
#     ws['H12'].alignment = Alignment(horizontal='right', vertical='center')

#     ws.merge_cells('I11:J11')
#     ws.merge_cells('I12:J12')
#     ws.merge_cells('I13:J13')
#     ws.merge_cells('I14:J14')
#     ws.merge_cells('I15:J15')



# # PI Date and PI Number
#     ws['I11'].value = pi.pi_date
#     ws['I12'].value = f'{pi.pi_no}'
#     date_style = NamedStyle(name="date_style", number_format="dd-mmm-yy")
#     ws['I11'].style = date_style
#     ws['I11'].font = Font(name='Calibri', size=9, bold=False)
#     ws['I12'].font = Font(name='Calibri', size=9, bold=False)
#     ws['I11'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
#     ws['I12'].alignment = Alignment(horizontal='left', vertical='center', indent=1)


# # For Tax Invoice and Date

#     ws['I13'].style = date_style
#     ws['I13'].font = Font(name='Calibri', size=9, bold=False)
#     ws['I14'].font = Font(name='Calibri', size=9, bold=False)
#     ws['I13'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
#     ws['I14'].alignment = Alignment(horizontal='left', vertical='center', indent=1)



# # For Contact Person and Team Member Name

#     ws.merge_cells('A19:C19')
#     ws.merge_cells('D19:F19')
#     ws.merge_cells('G19:H19')
#     ws.merge_cells('I19:J19')

#     ws['A19'].value = 'Team Member'
#     ws['D19'].value = 'Requistioner'
#     ws['G19'].value = 'Subscription Mode'
#     ws['I19'].value = 'Payment Term'

#     ws['A19'].fill = PatternFill(fill_type='solid', start_color="00C4BD97", end_color="00C4BD97")
#     ws['D19'].fill = PatternFill(fill_type='solid', start_color="00C4BD97", end_color="00C4BD97")
#     ws['G19'].fill = PatternFill(fill_type='solid', start_color="00C4BD97", end_color="00C4BD97")
#     ws['I19'].fill = PatternFill(fill_type='solid', start_color="00C4BD97", end_color="00C4BD97")

#     ws['A19'].font = Font(name='Calibri', bold=True, size=10)
#     ws['D19'].font = Font(name='Calibri', bold=True, size=10)
#     ws['G19'].font = Font(name='Calibri', bold=True, size=10)
#     ws['I19'].font = Font(name='Calibri', bold=True, size=10)

#     ws['A19'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['D19'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['G19'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['I19'].alignment = Alignment(horizontal='center', vertical='center')

#     thin_border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), vertical=Side(border_style='thin'),horizontal=Side(border_style='thin'))

#     for row in ws['A19:J20']:
#         for cell in row:
#             cell.border = thin_border

#     ws.merge_cells('A20:C20')
#     ws.merge_cells('D20:F20')
#     ws.merge_cells('G20:H20')
#     ws.merge_cells('I20:J20')

#     ws['A20'].value = f'{user_profile.user.first_name} {user_profile.user.last_name}'
#     ws['D20'].value = f'{pi.requistioner}'
#     ws['G20'].value = f'{pi.subscription}'
#     ws['I20'].value = f'{pi.payment_term}'

#     ws['A20'].font = Font(name='Calibri', bold=False, size=10)
#     ws['D20'].font = Font(name='Calibri', bold=False, size=10)
#     ws['G20'].font = Font(name='Calibri', bold=False, size=10)
#     ws['I20'].font = Font(name='Calibri', bold=False, size=10)

#     ws['A20'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['D20'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['G20'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['I20'].alignment = Alignment(horizontal='center', vertical='center')


# # Order Details Section Starts from here

#     ws.merge_cells('B22:G22')
#     ws.merge_cells('H22:I22')

#     ws['A22'].value = 'S.No.'
#     ws['B22'].value = 'Description'
#     ws['H22'].value = 'Unit Price '+f'({pi.currency})'.upper()
#     ws['J22'].value = 'Total '+f'({pi.currency})'.upper()

#     ws['A22'].font = Font(name='Calibri', bold=True, size=10)
#     ws['B22'].font = Font(name='Calibri', bold=True, size=10)
#     ws['H22'].font = Font(name='Calibri', bold=True, size=10)
#     ws['J22'].font = Font(name='Calibri', bold=True, size=10)

#     ws['A22'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['B22'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['H22'].alignment = Alignment(horizontal='center', vertical='center')
#     ws['J22'].alignment = Alignment(horizontal='center', vertical='center')

#     for row in ws['A22:J22']:
#         for cell in row:
#             cell.border = thin_border

#     n = 23
#     s = 1
#     for order in orders:
#         ws.merge_cells(f'A{n}:A{n+3}')
#         ws.merge_cells(f'H{n}:I{n+3}')
#         ws.merge_cells(f'J{n}:J{n+3}')

#         for row in ws[f'A{n}:J{n+3}']:
#             for cell in row:
#                 cell.border = thin_border

#         for row in ws[f'B{n}:G{n+3}']:
#             for cell in row:
#                 cell.border = None
            
#         for row in ws[f'B{n+3}:G{n+3}']:
#             for cell in row:
#                 cell.border = Border(bottom=Side(border_style='thin'))

#         ws[f'A{n}'].value = s
#         ws[f'A{n}'].font = Font(name='Calibri', bold=True, size=9)
#         ws[f'A{n}'].alignment = Alignment(horizontal='center', vertical='center')

#         cate = ws[f'B{n}']
#         cate.value = str(dict(CATEGORY).get(str(order.category), "Unknown Category"))
#         cate.font = Font(name='Calibri', size=10, bold=True)

#         sac = ws[f'G{n}']
#         sac.value = 'SAC: 998399'
#         sac.font = Font(name='Calibri', size=10, bold=True)
#         sac.alignment = Alignment(horizontal='right', vertical='bottom')

#         ws[f'B{n+1}'].value = 'News:'
#         ws[f'B{n+2}'].value = 'Product/HSN:'
#         ws[f'B{n+3}'].value = 'Period:'

#         from_month = datetime.strptime(order.from_month, '%Y-%m').strftime("%b'%y")
#         to_month = datetime.strptime(order.to_month, '%Y-%m').strftime("%b'%y")

#         ws[f'C{n+1}'].value = str(dict(REPORT_TYPE).get(str(order.report_type), "Unknown Report Type"))
#         ws[f'C{n+2}'].value = f'{order.product}'
#         ws[f'C{n+3}'].value = f'{from_month} - {to_month}'


#         for col in ws[f'C{n+1}:C{n+3}']:
#             for cell in col:
#                 cell.font = Font(name='Calibri', bold=True, size=10)

#         if pi.currency == 'inr':
#             ws[f'H{n}'].value = f'Rs. {order.unit_price}'
#             ws[f'J{n}'].value = f'{order.total_price:.2f}'
#         else:
#             ws[f'J{n}'].value = f'Usd. {order.unit_price}'
#             ws[f'J{n}'].value = f'{order.total_price:.2f}'
        
#         ws[f'H{n}'].font = Font(name='Calibri', bold=False, size=10)
#         ws[f'H{n}'].alignment = Alignment(horizontal='right', indent=1, vertical='center')
#         ws[f'J{n}'].font = Font(name='Calibri', bold=False, size=10)
#         ws[f'J{n}'].alignment = Alignment(horizontal='right', vertical='center')

#         for row in ws[f'B{n+1}:B{n+3}']:
#             for cell in row:
#                 cell.font = Font(name='Calibri', size=9)

#         n+= 4
#         s+= 1

#     ws.merge_cells(f'A{n}:B{n+1}')
#     ws.merge_cells(f'C{n}:G{n+1}')
#     ws.merge_cells(f'H{n}:I{n}')
#     ws.merge_cells(f'H{n+3}:I{n+3}')
#     ws.merge_cells(f'A{n+3}:G{n+8}')
#     ws.merge_cells(f'H{n+8}:J{n+11}')

#     for row in ws[f'A{n}:J{n+3}']:
#         for cell in row:
#             cell.border = thin_border

#     for row in ws[f'A{n+2}:G{n+3}']:
#         for cell in row:
#             cell.border = None

#     ws[f'H{n}'].value = 'Amount:'
#     ws[f'H{n}'].font = Font(name='Calibri', size=10, bold=True )
#     ws[f'H{n}'].alignment = Alignment(horizontal='right', indent=1, vertical='center')
#     ws[f'J{n}'].value = f'{net_total:.2f}'
#     ws[f'J{n}'].font = Font(name='Calibri', size=10, bold=True )
#     ws[f'J{n}'].alignment = Alignment(horizontal='right', vertical='center')

#     ws[f'A{n}'].value = CellRichText([TextBlock(InlineFont(b=True, i=False, sz=10), 'Amount Payable\n'), TextBlock(InlineFont(b=False, i=True, sz=8), '(in words)') ])
#     ws[f'A{n}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     if pi.currency == 'inr':
#         ws[f'C{n}'].value = 'Rs. '+ total_val_words + ' Only'
#     else:
#         ws[f'C{n}'].value = 'USD '+ total_val_words + ' Only'

#     ws[f'C{n}'].font =  Font(name='Calibri', size=10, bold=True )
#     ws[f'C{n}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

#     ws[f'A{n+3}'].value = CellRichText([TextBlock(normal_i_s,'Kindly pay in favor of\n'),TextBlock(bold_i_s,f'{pi.bank.bnf_name}\n'),
#                                         TextBlock(normal_i_s,f'Bank Name: {pi.bank.bank_name}\nA/C No.: {pi.bank.ac_no}\nBranch Address: {pi.bank.branch_address}\n'),
#                                         TextBlock(bold_i_s, f'IFSC: {pi.bank.ifsc}')])

#     ws[f'A{n+10}'].value = CellRichText([TextBlock(InlineFont(b=False, sz=9), 'Whether Tax is payable under REVERSE CHARGE:')])
#     ws[f'A{n+11}'].value = CellRichText([TextBlock(InlineFont(b=False, sz=9), 'Yes/No:')])
#     ws[f'B{n+11}'].value = CellRichText([TextBlock(InlineFont(b=True, sz=11), 'NO')])
#     ws[f'H{n+8}'].value = CellRichText([TextBlock(InlineFont(b=False, sz=9), 'Sign with Stamp')])


#     ws[f'A{n+3}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
#     ws[f'A{n+10}'].alignment = Alignment(horizontal='left', vertical='center')
#     ws[f'A{n+11}'].alignment = Alignment(horizontal='left', vertical='center')
#     ws[f'B{n+11}'].alignment = Alignment(horizontal='center', vertical='center')
#     ws[f'H{n+8}'].alignment = Alignment(horizontal='right', vertical='bottom')



# # Calculation of Tax Parts

#     if str(pi.state) == pi.bank.biller_id.biller_gstin[0:2]:    # Check tax calculation of CGST & SGST
#         ws.merge_cells(f'H{n+1}:I{n+1}')
#         ws.merge_cells(f'H{n+2}:I{n+2}')

#         ws[f'H{n+1}'].value = 'CGST(9%):'
#         ws[f'H{n+2}'].value = 'SGST(9%):'
#         ws[f'H{n+1}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'H{n+2}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'H{n+1}'].alignment = Alignment(horizontal='right', vertical='center', indent=1)
#         ws[f'H{n+2}'].alignment = Alignment(horizontal='right', vertical='center', indent=1)

#         ws[f'J{n+1}'].value = f'{cgst:.2f}'
#         ws[f'J{n+2}'].value = f'{sgst:.2f}'
#         ws[f'J{n+1}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'J{n+2}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'J{n+1}'].alignment = Alignment(horizontal='right', vertical='center')
#         ws[f'J{n+2}'].alignment = Alignment(horizontal='right', vertical='center')
#     else:
#         ws.merge_cells(f'H{n+1}:I{n+2}')
#         ws.merge_cells(f'J{n+1}:J{n+2}')
#         ws[f'H{n+1}'].value = 'IGST(18%):'
#         ws[f'H{n+1}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'H{n+1}'].alignment = Alignment(horizontal='right', vertical='center', indent=1)

#         ws[f'J{n+1}'].value = f'{igst:.2f}'
#         ws[f'J{n+1}'].font =  Font(name='Calibri', size=10, bold=False )
#         ws[f'J{n+1}'].alignment = Alignment(horizontal='right', vertical='center')

# # Total Value Including Tax
#     ws[f'H{n+3}'].value = 'TOTAL:'
#     ws[f'H{n+3}'].font =  Font(name='Calibri', size=11, bold=True )
#     ws[f'H{n+3}'].alignment = Alignment(horizontal='right', vertical='center', indent=1)
#     ws[f'J{n+3}'].value = f'{total_inc_tax:.2f}'
#     ws[f'J{n+3}'].font =  Font(name='Calibri', size=11, bold=True )
#     ws[f'J{n+3}'].alignment = Alignment(horizontal='right', vertical='center')


# # Pagesize Setup to Print
#     ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # Portrait orientation
#     ws.page_setup.paperSize = ws.PAPERSIZE_A4
#     ws.page_setup.fitToHeight = 1 
#     ws.page_setup.fitToWidth = 1
#     ws.page_margins = PageMargins(left=.55, right=0.5, top=0.75, bottom=0.75)

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename=PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.xlsx'

#     # Save the workbook to the response object
#     wb.save(response)

#     return response



# @login_required(login_url='app:login')
# def download_pdf(request, pi_id):
#     pi = get_object_or_404(proforma, pk = pi_id)
#     user = Profile.objects.get(user=pi.user_id)
#     orders = orderList.objects.filter(proforma_id= pi.id)

#     reg_address = pi.bank.biller_id.get_reg_full_address()
#     corp_address = pi.bank.biller_id.get_corp_full_address()

#     context = {
#         'pi': pi,
#         'user': user,
#     }

#     buffer = BytesIO()

#     doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.2 * inch, bottomMargin=0.2 * inch,)

#     blue_font = ParagraphStyle(name="BlueStyle", fontSize=24, alignment=2, fontName='Helvetica-Bold', underlineWidth=2 )
#     blue_font_m = ParagraphStyle(name="BlueStyle", fontSize=20, alignment=2, fontName='Helvetica-Bold', underlineWidth=2, leading = 24 )

#     brand_text = f'{pi.bank.biller_id.biller_name}'
#     if pi.bank.biller_id.biller_gstin != '':
#         brand_name = Paragraph('<u><font color="#00CCFF">BE </font><u><font color="orange">SMART EXIM</font></u></u>', blue_font)
#     else:
#         brand_name = Paragraph(f'<u><font color="#00CCFF">{brand_text}</font></u>', blue_font_m)


#     biller_font = ParagraphStyle(name="BlueStyle", fontSize=8, alignment=0, fontName='Helvetica', underlineWidth=2 )

#     corp_add_text = f'<br/><b>Corp. Office: </b><font>{corp_address}</font>' if corp_address else ''
#     gstin_text = f'<br/><b>GSTIN: </b><font>{pi.bank.biller_id.biller_gstin}</font>' if corp_address else ''

#     biller_dtl = Paragraph(f"<i><b>Issued by:</b><br/><b>{pi.bank.biller_id.biller_name}</b><br/><b>Reg. Office: </b><font>{reg_address}</font>{corp_add_text} {gstin_text}></i>", biller_font)

#     user_dtl = None
#     if pi.bank.biller_id.biller_gstin != '':
#         user_dtl = Paragraph(f"<b>Email: </b><font>{user.user.email} || </font><b>Contact: </b><font>{user.phone}</font><br/><b>Website: </b><font>www.besmartexim.com</font>", biller_font)

#     font_m = ParagraphStyle(name="BlueStyle", fontSize=10, fontName='Helvetica', wordWrap=False)
#     font_m_c = ParagraphStyle(name="BlueStyle", fontSize=10, fontName='Helvetica', wordWrap=False, alignment=1)
#     font_s = ParagraphStyle(name="BlueStyle", fontSize=9, fontName='Helvetica', wordWrap=False)
#     font_s_c = ParagraphStyle(name="BlueStyle", fontSize=9, fontName='Helvetica', wordWrap=False, alignment=1)
#     company = Paragraph(f"<b>{pi.company_name}</b>", font_m)
#     gstin = Paragraph(f"<b>GSTIN: </b><font>{pi.gstin}</font>", font_s)
#     address = Paragraph(f"<b>Address: </b><font>{pi.address}</font>", font_s)
#     email = Paragraph(f"<b>Email: </b><font>{pi.email_id}</font>", font_s)
#     contact = Paragraph(f"<b>Contact: </b><font>{pi.contact}</font>", font_s)

#     piDate = pi.pi_date.strftime("%d-%b-%y")

#     user_name = Paragraph(f"<font>{user.user.first_name} {user.user.last_name}</font>", font_m_c)
#     requistioner = Paragraph(f"<font>{pi.requistioner}</font>", font_m_c)
#     subs = Paragraph(f"<font>{pi.subscription}</font>", font_m_c)
#     pay_m = Paragraph(f"<font>{pi.payment_term}</font>", font_m_c)

#     unit_price_h = Paragraph(f"<b><font>Unit Price ({pi.currency.upper()})</font></b>", font_s_c)
#     total_price_h = Paragraph(f"<b><font>Total ({pi.currency.upper()})</font></b>", font_s_c)

#     basedir = Path(__file__).resolve().parent.parent


#     imagepath = os.path.join(basedir,'static/becrm/image/PI_LOGO.png')
#     logo = Im(imagepath,0.85*inch,0.85*inch)

#     data = [
#         ['','','','','','PRO FORMA INVOICE','','','',''],
#         ['','','','','','','','','',''],
#         [logo,'','',brand_name,'','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','founded by SUN TRADE INC','','','','',],
#         [biller_dtl,'','','','',user_dtl,'','','','',],
#         ['','','','','','','','','',''],
#         ['Customer Details:','','','','','','','','',''],
#         [company,'','','','','','','PI DATE:',piDate,''],
#         [gstin,'','','','','','','PI NUMBER:',pi.pi_no,''],
#         [address,'','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         [email,'','','','','','','','',''],
#         [contact,'','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['Team Member','','','Requistioner','','','Subscription Mode','','Payment Term',''],
#         [user_name,'','',requistioner,'','',subs,'',pay_m,''],
#         ['','','','','','','','','',''],
#         ['S.No.','Description','','','','','',unit_price_h,'',total_price_h],
#     ]


#     row_heights = [0.4*inch, 0.1*inch] + 3*[0.4*inch] + [1.2*inch] + 12*[0.2*inch] + [0.1*inch, 0.2*inch]
#     col_widths = [0.7*inch, 0.85*inch] + 6*[0.7*inch] + [0.85*inch]

#     table = Table(data, col_widths, row_heights)

#     style = TableStyle([
#         ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#         ('SPAN',(5,0),(9,0)),
#         ('ALIGN', (5, 0), (9, 4), 'RIGHT'),
#         ('TEXTCOLOR', (0, 0), (9, 0), colors.HexColor("#948A54")),
#         ('FONTSIZE', (0, 0), (9, 0), 21),
#         ('SPAN',(3,2),(9,3)),
#         ('BACKGROUND', (0, 1), (2, 1), colors.HexColor("#5097D7")),
#         ('BACKGROUND', (3, 1), (9, 1), colors.HexColor("#E0A800")),
#         ('SPAN',(5,4),(9,4)),
#         ('TEXTCOLOR', (5,4),(9,4), colors.HexColor("#244062")),
#         ('FONT', (5,4),(9,4), 'Helvetica-Bold', 9),
#         ('VALIGN', (0, 0), (9, 5), 'TOP'),
#         ('VALIGN', (3, 2), (9, 4), 'MIDDLE'),
#         ('VALIGN', (0, 5), (-1, -1), 'MIDDLE'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('SPAN',(0,5),(4,5)),
#         ('SPAN',(5,5),(9,5)),
#         ('LINEBEFORE',(5,5),(9,5),0.5,colors.black),
#         # ('LINEABOVE',(1,6),(8,6),0.5,colors.black),
#         ('SPAN',(5,5),(9,5)),
#         ('SPAN',(0,8),(4,8)),
#         ('SPAN',(0,9),(4,9)),
#         ('SPAN',(0,10),(4,12)),
#         ('VALIGN',(0,10),(4,12), 'TOP'),
#         ('SPAN',(0,13),(4,13)),
#         ('SPAN',(0,14),(4,14)),
#         ('ALIGN', (5, 8), (7, 13), 'RIGHT'),
#         ('FONTSIZE', (5, 8), (7, 13), 8),
#         ('FONTSIZE', (8, 8), (9, 13), 8),
#         ('VALIGN', (5, 8), (7, 13), 'MIDDLE'),
#         ('SPAN',(8,8),(9,8)),
#         ('SPAN',(8,9),(9,9)),
#         ('SPAN',(8,10),(9,10)),
#         ('SPAN',(8,11),(9,11)),
#         ('SPAN',(8,12),(9,12)),
#         ('SPAN',(8,13),(9,13)),
#         ('SPAN',(0,16),(2,16)),
#         ('SPAN',(0,17),(2,17)),
#         ('SPAN',(3,16),(5,16)),
#         ('SPAN',(3,17),(5,17)),
#         ('SPAN',(6,16),(7,16)),
#         ('SPAN',(6,17),(7,17)),
#         ('SPAN',(8,16),(9,16)),
#         ('SPAN',(8,17),(9,17)),
#         ('BACKGROUND',(0,16),(9,16), colors.HexColor("#C4BD97")),
#         ('ALIGN',(0,16),(9,19), 'CENTER'),
#         ('FONT',(0,16),(9,16), 'Helvetica-Bold', 9),
#         ('FONT',(0,17),(9,17), 'Helvetica', 9),
#         ('GRID',(0,16),(9,17),0.5,colors.black),
#         ('SPAN',(1,19),(6,19)),
#         ('SPAN',(7,19),(8,19)),
#         ('FONT',(0,19),(9,19), 'Helvetica-Bold', 9),
#         ('GRID',(0,19),(9,19),0.5,colors.black)
#     ])

#     table.setStyle(style)

#     nr = 0
#     s = 1
#     order_list = []
#     lumpsumRow = 0

#     net_total = total_order_value(pi)
#     lumpsumAmt = total_lumpsums(pi)

#     for order in filter_by_lumpsum(orders,True):
#         if order.is_lumpsum:
#             cat = dict(CATEGORY).get(str(order.category))
#             report = dict(REPORT_TYPE).get(str(order.report_type))
#             product = order.product
#             from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
#             to_month = datetime.strptime(order.to_month, '%Y-%m').strftime("%b'%y")
#             period = f'{from_month} - {to_month}'

#             if pi.currency == 'inr':
#                 unitPrice = f'Lumpsum'
#                 totalPrice = f'{lumpsumAmt:.2f}'
#             else:
#                 unitPrice = f'Lumpsum'
#                 totalPrice = f'{lumpsumAmt:.2f}'

#             sac = f'SAC: 998399'

#         from_m = datetime.strptime(order.from_month, '%Y-%m')
#         to_m = datetime.strptime(order.to_month, '%Y-%m')

#         no_months = (to_m.year - from_m.year)*12 + to_m.month - from_m.month + 1

#         total_m = f'Total: {no_months}'

#         order_list.extend([
#             [s,cat,'','','','',sac,'','',''],
#             ['','News:',report,'','','','','','',''],
#             ['','Product/HSN:',product,'','','','','','',''],
#             ['','Period:',period,'','','',total_m,'','',''],
#             ])
#         nr+= 4
#         s+= 1
#         lumpsumRow+=4

#     for order in filter_by_lumpsum(orders,False):
#             cat = dict(CATEGORY).get(str(order.category))
#             report = dict(REPORT_TYPE).get(str(order.report_type))
#             product = order.product
#             from_month = datetime.strptime(order.from_month, "%Y-%m").strftime("%b'%y")
#             to_month = datetime.strptime(order.to_month, '%Y-%m').strftime("%b'%y")
#             period = f'{from_month} - {to_month}'

#             if pi.currency == 'inr':
#                 unitPrice = f'Rs. {order.unit_price}'
#                 totalPrice = f'{order.total_price:.2f}'
#             else:
#                 unitPrice = f'Usd. {order.unit_price}'
#                 totalPrice = f'{order.total_price:.2f}'

#             sac = f'SAC: 998399'

#             from_m = datetime.strptime(order.from_month, '%Y-%m')
#             to_m = datetime.strptime(order.to_month, '%Y-%m')

#             no_months = (to_m.year - from_m.year)*12 + to_m.month - from_m.month + 1

#             total_m = f'Total: {no_months}'

#             order_list.extend([
#                 [s,cat,'','','','',sac,unitPrice,'',totalPrice],
#                 ['','News:',report,'','','','','','',''],
#                 ['','Product/HSN:',product,'','','','','','',''],
#                 ['','Period:',period,'','','',total_m,'','',''],
#                 ])
#             nr+= 4
#             s+= 1


#     words_amt = Paragraph(f"<b>Amount Payable</b><br/><i><font size='8'>(in words)</font></i>", font_s_c)

#     if pi.is_sez:
#         cgst = 0
#         sgst = 0
#         igst = 0
#         total_inc_tax = net_total
#     else:
#         cgst = net_total*0.09
#         sgst = net_total*0.09
#         igst = net_total*0.18
#         total_inc_tax = net_total*1.18 

#     if pi.currency == 'inr':
#         total_val_words = Paragraph(f'<b><font>Rs. {num2words(total_inc_tax, lang='en').replace(',', '').title()} Only</font></b>', font_s)
#     else:
#         total_val_words = Paragraph(f'<b><font>Usd {num2words(total_inc_tax, lang='en').replace(',', '').title()} Only</font></b>', font_s)


#     bank_dtl = Paragraph(f"<i><font>Kindly pay in favor of</font><br/><b>{pi.bank.bnf_name}</b><br/><font>Bank Name: {pi.bank.bank_name}</font><br/><font>A/C No.: {pi.bank.ac_no}</font><br/><font>Branch Address: {pi.bank.branch_address}</font><br/><b>IFSC: {pi.bank.ifsc}</b></i>", font_s)


#     order_list.extend([
#         [words_amt,'',total_val_words,'','','','','Amount:','',f'{net_total:.2f}'],
#     ])

#     if str(pi.state) == pi.bank.biller_id.biller_gstin[0:2]:

#             order_list.extend([
#                 ['','','','','','','','CGST(9%):','',f'{cgst:.2f}'],
#                 ['','','','','','','','SGST(9%):','',f'{sgst:.2f}'],
#             ])
#     else:

#             order_list.extend([
#                 ['','','','','','','','IGST(18%):','',f'{igst:.2f}'],
#                 ['','','','','','','','','',''],
#             ])



#     order_list.extend([
#         [bank_dtl,'','','','','','','TOTAL:','',f'{total_inc_tax:.2f}'],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['','','','','','','','','',''],
#         ['Whether Tax is payable under REVERSE CHARGE:','','','','','','','','',''],
#         ['Yes/No:','NO','','','','','','Sign and Stamp','',''],
#     ])


#     order_style = []

#     if lumpsumRow>0:
#         for r in range(0, lumpsumRow, 4):
#             end_row = min(lumpsumRow - 1, r + 3)
#             order_list[0][7] = "Lumpsum"
#             order_list[0][9] = f'{lumpsumAmt:.2f}'
#             for i in range(r + 1, end_row + 1):
#                 order_list[i][7] = ''
#                 order_list[i][8] = ''
#                 order_list[i][9] = ''
#             order_style.extend([
#                 ('SPAN',(7,r),(8, lumpsumRow-1)),
#                 ('SPAN',(9,r),(9, lumpsumRow-1)),])
        
#     if nr>lumpsumRow:
#         for r in range(lumpsumRow,nr, 4):
#             order_style.extend([
#                 ('SPAN',(7,r),(8, r+3)),
#                 ('SPAN',(9,r),(9, r+3)),])

#     for r in range(0, nr, 4):
#         order_style.extend([
#             ('SPAN',(0,r),(0, r+3)),
#             ('FONT',(1,r),(6, r),'Helvetica-Bold',10),
#             ('FONT',(1,r+1),(1, r+3),'Helvetica',8),
#             ('VALIGN',(0,r),(0, r+3), 'MIDDLE'),
#             ('ALIGN',(0,r),(0, r+3), 'CENTER'),
#             ('GRID',(0,r),(0,r+3),0.5,colors.black),
#             ('FONTSIZE',(6,r+1),(6, r+3), 8),
#             ('RIGHTPADDING',(6,r),(6, r+3), 2),
#             ('LINEBELOW',(1,r+3),(6,r+3),0.5,colors.black),
#             ('ALIGN',(6,r),(9, nr+3), 'RIGHT'),
#             ('VALIGN',(6,r),(9, nr+3), 'MIDDLE'),
#             ('VALIGN',(0,nr),(9, nr+3), 'MIDDLE'),
#             ('FONTNAME',(6,nr),(9, nr), 'Helvetica-Bold'),
#             ('FONTNAME',(6,nr+3),(9, nr+3), 'Helvetica-Bold'),
#             ('FONTSIZE',(6,nr+1),(9, nr+2), 9),
#             ('RIGHTPADDING',(7,r),(8, nr+3), 10),
#             ('GRID',(7,r),(8,r+3),0.5,colors.black),
#             ('GRID',(9,r),(9,r+3),0.5,colors.black),
#             ('SPAN',(0,nr),(1, nr+1)),
#             ('SPAN',(2,nr),(6, nr+1)),
#             ('SPAN',(7,nr),(8, nr)),
#             ('SPAN',(7,nr+1),(8, nr+1)),
#             ('SPAN',(7,nr+2),(8, nr+2)),
#             ('SPAN',(7,nr+3),(8, nr+3)),
#             ('SPAN',(0,nr+3),(6, nr+8)),
#             ('SPAN',(7,nr+11),(9, nr+11)),
#             ('FONTSIZE',(0,nr+10),(0, nr+11), 8),
#             ('ALIGN',(0,nr+10),(0, nr+11), 'LEFT'),
#             ('VALIGN',(0,nr+10),(0, nr+11), 'MIDDLE'),
#             ('ALIGN',(1,nr+11),(1, nr+11), 'CENTER'),
#             ('FONT',(1,nr+11),(1, nr+11), 'Helvetica-Bold'),
#             ('ALIGN',(7,nr+11),(9, nr+11), 'RIGHT'),
#             ('GRID',(0,nr),(9,nr+1),0.5,colors.black),
#             ('GRID',(7,nr),(9,nr+3),0.5,colors.black),
#             # ('GRID',(0,nr),(9,nr+11),0.5,colors.black),
#         ])

#     if str(pi.state) != pi.bank.biller_id.biller_gstin[0:2]:
#         order_style.append(('SPAN',(7,nr+1), (8,nr+2)))
#         order_style.append(('SPAN',(9,nr+1), (9,nr+2)))


#     order_table = Table(order_list, col_widths, (nr+12)*[0.21*inch])

#     order_table.setStyle(TableStyle(order_style))

#     elements = []
#     elements.append(table)
#     elements.append(order_table)

#     doc.build(elements)

#     buffer.seek(0)

#     response = HttpResponse(buffer.getvalue(), content_type = 'application/pdf')

#     response['Content-Disposition'] = f'attachment; filename=PI_{pi.company_name}_{pi.pi_no}_{pi.pi_date}.pdf'   

#     return response