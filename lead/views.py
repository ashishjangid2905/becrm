# Import from Django
from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.db.models import Subquery, Max, F, Q
from django.conf import settings
from django.db import transaction, IntegrityError

# Import from  app models, utils
from .models import leads, contactPerson, Conversation, conversationDetails
from teams.models import User, Profile
from teams.templatetags.teams_custom_filters import get_current_position

# Import Third Party Module
import csv, os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import logging
from django.utils.timezone import now

# Create your views here.

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from .serializers import *
from rest_framework.pagination import PageNumberPagination
from rest_framework.filters import OrderingFilter, SearchFilter
from django_filters.rest_framework import FilterSet
from django.db import IntegrityError, transaction


class BasePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class LeadsFilters(FilterSet):
    class Meta:
        model = leads
        fields = ["company_name", "gstin", "full_address", "industry"]


class lead_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = BasePagination
    search_fields = ["company_name", "gstin", "full_address", "industry"]
    ordering_fields = [
        "company_name",
        "gstin",
        "full_address",
        "industry",
        "user_name",
        "created_at",
    ]
    ordering = ["-created_at"]

    def get(self, request):
        try:
            all_leads = leads.objects.filter(user=request.user.id).prefetch_related(
                "contactpersons"
            )
            search_query = request.GET.get("search", None)
            if search_query:
                search_filter = SearchFilter()
                all_leads = search_filter.filter_queryset(request, all_leads, self)

            filtered_leads = LeadsFilters(request.GET, queryset=all_leads)
            if filtered_leads.is_valid():
                all_leads = filtered_leads.qs

            ordering_query = request.GET.get("ordering", None)
            if ordering_query:
                ordering_filter = OrderingFilter()
                all_leads = ordering_filter.filter_queryset(request, all_leads, self)

            paginator = self.pagination_class()
            paginated_leads = paginator.paginate_queryset(all_leads, request)
            serializer = leadsSerializer(paginated_leads, many=True)
            return paginator.get_paginated_response(serializer.data)
        except leads.DoesNotExist:
            return Response(
                {"message": "No Lead founded"}, status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        serializer = leadsSerializer(data=request.data)
        user_id = request.user.id
        try:
            if serializer.is_valid():
                company_name = request.data.get("company_name")
                gstin = request.data.get("gstin")
                if leads.objects.filter(
                    company_name=company_name, gstin=gstin, user=user_id
                ).exists():
                    return Response(
                        {"error": f"The Company {company_name} already exists."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                with transaction.atomic():
                    lead_instance = serializer.save(user=user_id)
                    contactpersons_data = request.data.get("contactpersons", [])
                    lead = get_object_or_404(leads, pk=lead_instance.id)
                    contact_persons = [
                        contactPerson(company=lead, **contact)
                        for contact in contactpersons_data
                    ]
                    if contact_persons:
                        contactPerson.objects.bulk_create(contact_persons)
                    return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(
                {"error": "Field Vaildation Error", "details": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except IntegrityError as ie:
            return Response(
                {"error": "Database integrity error", "details": str(ie)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class LeadListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            all_leads = (
                leads.objects.filter(user=request.user.id)
                .prefetch_related("contactpersons")
                .order_by("company_name")
            )

            serializer = leadsSerializer(all_leads, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except leads.DoesNotExist:
            return Response(
                {"message": "No Lead founded"}, status=status.HTTP_404_NOT_FOUND
            )


class LeadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        try:
            lead = get_object_or_404(leads, pk=id)
            serializer = leadsSerializer(lead)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, id):
        try:
            instance = get_object_or_404(leads, pk=id)
            data = request.data
            serializer = leadsSerializer(instance, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ContactView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, id):
        try:
            leadInstance = get_object_or_404(leads, id=id)
            data = request.data
            company_id = data.pop("company")
            print(data)
            serializer = contactPersonSerializer(data=data)
            if serializer.is_valid():
                serializer.save(company=leadInstance)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, lead_id, id):
        try:
            instance = get_object_or_404(contactPerson, id=id)
            data = request.data
            data.pop("id")
            serializer = contactPersonSerializer(instance, data=data, partial=True)

            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ConversationView(APIView):
    permission_classes = [IsAuthenticated]

    # get conversation object
    def get_object(self, id):
        return get_object_or_404(Conversation, id=id)

    # fetch conversation list with respect to lead
    def get(self, request, id):
        try:
            conversation = (
                Conversation.objects.filter(company_id=id)
                .select_related("company_id")
                .order_by("-start_at")
            )
            serializer = ConversationSerializer(conversation, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Conversation.DoesNotExist:
            return Response(
                {"message": "No Conversation found for this Company"},
                status=status.HTTP_404_NOT_FOUND,
            )

    # Create/Start New Conversation
    def post(self, request, id):
        try:
            data = request.data.copy()
            title = data.get("title")
            company_id = data.get("company")
            if not title or not company_id:
                return Response(
                    {"error": "Missing 'title' or 'company' in request data"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                deal_data = {"title": title, "company_id": company_id}
                dealSerializer = ConversationSerializer(data=deal_data)
                if not dealSerializer.is_valid():
                    raise IntegrityError(
                        dealSerializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
                deal = dealSerializer.save()

                serializer = ConversationDetailsSerializer(data=data)
                if not serializer.is_valid():
                    raise IntegrityError(
                        serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )

                serializer.save(chat_no=deal)
                return Response(dealSerializer.data, status=status.HTTP_201_CREATED)

        except IntegrityError as ie:
            return Response({"error": str(ie)}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class dealActivityView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, id):
        return get_object_or_404(Conversation, id=id)

    def get(self, request, id):
        try:
            deal = self.get_object(id)

            activities = conversationDetails.objects.filter(
                chat_no=deal
            ).select_related("chat_no")
            if activities:
                serializer = ConversationDetailsSerializer(activities, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response({}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request, id):
        try:
            deal = self.get_object(id)
            # print(deal)
            data = request.data
            chat_no = data.pop("chat_no")
            if data.get("follow_up") == "":
                data["follow_up"] = None

            print(data)

            serializer = ConversationDetailsSerializer(data=data)

            if serializer.is_valid():
                serializer.save(chat_no=deal)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class InboundLeadGetView(APIView):
    permission_classes = [AllowAny]
    allowed_origins = [
        "https://www.besmartexim.com",
        "https://www.bedatos.com",
        "http://192.168.3.98:8000",
        "http://localhost:3000",  # if testing locally
    ]

    def post(self, request):
        try:
            origin = request.headers.get("Origin") or request.headers.get("Referer")
            print(origin)
            if origin and not any(
                origin.startswith(allowed) for allowed in self.allowed_origins
            ):
                return Response(
                    {"error": "Unauthorized origin"},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
            data = request.data
            serializer = InboundLeadSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(
                    {"message": "Thank you for the subscription"},
                    status=status.HTTP_201_CREATED,
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url="app:login")
def follow_ups(request):

    user_id = request.user.id

    profile_instance = get_object_or_404(Profile, user=request.user)
    user_branch = profile_instance.branch.id
    all_users = Profile.objects.filter(branch=user_branch)

    users_ids = [user.user.id for user in all_users]

    if request.user.role == "admin":
        leads_ids = leads.objects.filter(user__in=users_ids).values_list(
            "id", flat=True
        )
    else:
        leads_ids = leads.objects.filter(user=user_id).values_list("id", flat=True)

    conversations_ids = Conversation.objects.filter(
        company_id__in=leads_ids
    ).values_list("id", flat=True)

    allchat = conversationDetails.objects.filter(
        chat_no__in=conversations_ids
    ).select_related("chat_no__company_id", "contact_person")

    latest_inserted = allchat.values("chat_no").annotate(
        latest_insert=Max("inserted_at")
    )

    allchat = allchat.filter(
        inserted_at__in=Subquery(latest_inserted.values("latest_insert"))
    )

    today = now().date()

    filter_value = request.GET.get("filter", "today")

    if filter_value == "today":
        filtered_conversations = allchat.filter(follow_up=today)
    elif filter_value == "previous":
        filtered_conversations = allchat.filter(follow_up__lt=today)
    elif filter_value == "future":
        filtered_conversations = allchat.filter(follow_up__gt=today)
    else:
        filtered_conversations = allchat.filter(follow_up=today)

    if filtered_conversations:
        for lead in filtered_conversations:
            lead.chat_no.company_id.user = get_object_or_404(
                User, pk=lead.chat_no.company_id.user
            )

    context = {
        "latest_conversation_details": filtered_conversations,
    }

    return render(request, "lead/follow-up-list.html", context)


@login_required(login_url="app:login")
def upload_Leads(request):

    if not request.user.is_authenticated:
        return redirect("app:login")
    logger = logging.getLogger(__name__)
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]
        if uploaded_file.name.endswith(".csv"):
            try:
                decoded_file = uploaded_file.read().decode("utf-8").splitlines()
                csv_reader = csv.DictReader(decoded_file)
            except Exception as e:
                logger.error(f"Error reading file: {str(e)}")
                return HttpResponse(f"Error reading file: {e}")
            # Handle potential BOM (Byte Order Mark) in the first column name
            fieldnames = csv_reader.fieldnames
            if fieldnames[0].startswith("\ufeff"):
                fieldnames[0] = fieldnames[0].replace("\ufeff", "")

            # Ensure the CSV DictReader knows about the corrected fieldnames
            csv_reader.fieldnames = fieldnames

            for row in csv_reader:
                try:
                    # Check if lead already exists based on company name and GSTIN
                    lead_instance, created = leads.objects.update_or_create(
                        company_name=row["company_name"],
                        gstin=row["gstin"],
                        user=request.user.id,
                        defaults={
                            "address1": row["address1"],
                            "address2": row.get("address2", ""),
                            "city": row["city"],
                            "state": row["state"],
                            "country": row["country"],
                            "pincode": row.get("pincode", ""),
                            "industry": row.get("industry", ""),
                            "source": row["source"],
                        },
                    )

                    # Update or create contact person associated with the lead
                    contact_person_instance, contact_created = (
                        contactPerson.objects.update_or_create(
                            person_name=row["person_name"],
                            email_id=row.get("email_id", ""),
                            contact_no=row.get("contact_no", ""),
                            company=lead_instance,
                            defaults={
                                "is_active": row.get("is_active", "TRUE").lower()
                                == "true"
                            },
                        )
                    )
                except Exception as e:
                    # Log the error or handle it
                    print(f"Error processing row: {row}. Error: {str(e)}")
                    continue  # Skip to the next row

            return redirect("lead:leads_list")
        else:
            return HttpResponse("Invalid file format")


@login_required(login_url="app:login")
def download_template(request):
    file_path = os.path.join(settings.BASE_DIR, "static", "becrm/template.csv")
    return FileResponse(
        open(file_path, "rb"), as_attachment=True, filename="template.csv"
    )


def exportlead(all_lead):

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Team Member",
        "Company Name",
        "GSTIN",
        "Address",
        "City",
        "State",
        "Country",
        "Pincode",
        "Industry",
        "Source",
        "Contact Person Name",
        "Email",
        "Contact Number",
        "Is Active",
    ]

    ws.append(headers)

    header_fill = PatternFill(
        fill_type="solid", start_color="0099CCFF", end_color="0099CCFF"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="00000000")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows to the sheet
    for lead in all_lead:
        # Iterate over each contact person associated with the lead
        contact_person = contactPerson.objects.filter(company=lead)

        for contact_person in contact_person:
            user_name = User.objects.get(pk=contact_person.company.user).get_full_name()
            address = contact_person.company.get_full_address()
            row = [
                user_name,
                contact_person.company.company_name,
                contact_person.company.gstin,
                address,
                contact_person.company.city,
                contact_person.company.state,
                contact_person.company.country,
                contact_person.company.pincode,
                contact_person.company.industry,
                contact_person.company.source,
                contact_person.person_name,
                contact_person.email_id,
                contact_person.contact_no,
                contact_person.is_active,
            ]
            ws.append(row)

    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Create a response object with MIME type for Excel files
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=leads.xlsx"

    # Save the workbook to the response object
    wb.save(response)

    return response
